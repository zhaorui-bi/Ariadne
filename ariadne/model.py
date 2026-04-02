"""ESM embedding workflow for coral TPS type separation from spreadsheet data.

This module is intentionally kept separate from the main four-stage Ariadne
pipeline because it targets a supervised analysis setting: a curated Excel file
containing known coral TPS proteins and their type labels. The typical use case
is to embed these proteins with ESM2, then evaluate how well the resulting
representation separates product types such as CeeSs-related classes and other
TPS groups.
"""

from __future__ import annotations

import logging
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Optional, Union

import numpy as np
from sklearn.base import BaseEstimator, ClassifierMixin, clone
from sklearn.decomposition import PCA
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, classification_report, confusion_matrix, f1_score
from sklearn.model_selection import StratifiedKFold, cross_val_predict
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler

from ariadne.utils import FastaRecord, clean_sequence, ensure_directory, write_fasta, write_tsv

PathLike = Union[str, Path]

logger = logging.getLogger(__name__)

try:
    from tqdm.auto import tqdm
except ImportError:  # pragma: no cover - fallback for minimal environments
    def tqdm(iterable, **kwargs):
        return iterable

DEFAULT_ESM_MODEL_NAME = "facebook/esm2_t33_650M_UR50D"
ESM_MODEL_PRESETS = {
    "150M": "facebook/esm2_t30_150M_UR50D",
    "650M": "facebook/esm2_t33_650M_UR50D",
    # Keep a compatibility shorthand for project users who refer to the larger checkpoint as "504M".
    "504M": "facebook/esm2_t33_650M_UR50D",
    "3B": "facebook/esm2_t36_3B_UR50D",
    "15B": "facebook/esm2_t48_15B_UR50D",
}
CEESS_POSITIVE_LABELS = frozenset({"cembrene A", "cembrene B"})


def resolve_esm_model_name(model_name: str) -> str:
    """Resolve user-facing ESM presets to concrete Hugging Face model ids."""
    key = str(model_name).strip()
    preset = ESM_MODEL_PRESETS.get(key.upper())
    if preset is not None:
        if key.upper() == "504M":
            logger.info("Requested ESM preset '504M'; using supported checkpoint %s", preset)
        return preset
    return key


def esm_model_help_text() -> str:
    """Return concise CLI help text listing common supported ESM presets."""
    return (
        "ESM model preset or Hugging Face id. "
        "Common presets: 150M, 650M, 3B, 15B. "
        "Default: 650M. The shorthand 504M is accepted and mapped to the bundled large ESM2 checkpoint."
    )


@dataclass
class TPSTypeRecord:
    """One labeled protein record loaded from ``TPS.xlsx``."""

    name: str
    sequence: str
    label: str
    species: str
    ceess_group: str = "non-CeeSs"


@dataclass
class CeessPredictionResult:
    """Outputs produced when scoring coral-like candidates for CeeSs."""

    output_paths: dict[str, Path]
    prediction_rows: list[dict[str, object]]
    candidate_ids: list[str]


def load_tps_xlsx(
    xlsx_path: PathLike,
    *,
    sheet_name: Optional[str] = None,
    name_column: str = "Name",
    sequence_column: str = "Protein",
    label_column: str = "Type",
    species_column: str = "Species",
    ceess_group_column: Optional[str] = None,
) -> list[TPSTypeRecord]:
    """Load coral TPS records from an Excel workbook.

    The expected spreadsheet layout follows the current `TPS/TPS.xlsx`
    convention:
    - column 1: name
    - column 2: protein sequence
    - column 3: type label
    - column 4: species
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read TPS.xlsx files.") from exc

    workbook = load_workbook(Path(xlsx_path), read_only=True, data_only=True)
    target_sheet = sheet_name or workbook.sheetnames[0]
    worksheet = workbook[target_sheet]
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        raise ValueError(f"Worksheet '{target_sheet}' is empty in {xlsx_path}.")

    header_map = {str(value).strip(): index for index, value in enumerate(header) if value is not None}
    required = {
        "name": name_column,
        "sequence": sequence_column,
        "label": label_column,
    }
    for field, column_name in required.items():
        if column_name not in header_map:
            raise ValueError(
                f"Missing required column '{column_name}' in worksheet '{target_sheet}'. "
                f"Available columns: {list(header_map)}"
            )

    candidate_ceess_group_columns = [
        column_name
        for column_name in (
            ceess_group_column,
            "CeeSs_group",
            "CeeSs Group",
            "ceess_group",
            "ceess group",
            "CeeSs",
            "ceess",
        )
        if column_name
    ]
    resolved_ceess_group_column = next((name for name in candidate_ceess_group_columns if name in header_map), None)

    def normalize_ceess_group(raw_value: object, label: str) -> str:
        if raw_value is None or str(raw_value).strip() == "":
            return "CeeSs" if str(label).strip() in CEESS_POSITIVE_LABELS else "non-CeeSs"
        token = str(raw_value).strip().lower().replace(" ", "_").replace("-", "_")
        if token in {"ceess", "yes", "y", "true", "1", "positive", "pos"}:
            return "CeeSs"
        if token in {"non_ceess", "no", "n", "false", "0", "negative", "neg"}:
            return "non-CeeSs"
        if "ceess" in token and "non" not in token:
            return "CeeSs"
        if "non" in token and "ceess" in token:
            return "non-CeeSs"
        logger.warning(
            "Unrecognized CeeSs-group value '%s' in %s; falling back to label-derived grouping for %s.",
            raw_value,
            xlsx_path,
            label,
        )
        return "CeeSs" if str(label).strip() in CEESS_POSITIVE_LABELS else "non-CeeSs"

    records: list[TPSTypeRecord] = []
    for row_index, row in enumerate(rows, start=2):
        name_value = row[header_map[name_column]]
        sequence_value = row[header_map[sequence_column]]
        label_value = row[header_map[label_column]]
        species_value = row[header_map[species_column]] if species_column in header_map else ""

        if not sequence_value or not label_value:
            continue
        sequence = clean_sequence(str(sequence_value), keep_gaps=False).replace("*", "")
        if not sequence:
            logger.warning("Skipped empty/invalid protein sequence at row %d in %s", row_index, xlsx_path)
            continue

        records.append(
            TPSTypeRecord(
                name=str(name_value).strip() if name_value else f"record_{row_index}",
                sequence=sequence,
                label=str(label_value).strip(),
                species=str(species_value).strip() if species_value else "",
                ceess_group=normalize_ceess_group(
                    row[header_map[resolved_ceess_group_column]] if resolved_ceess_group_column is not None else None,
                    str(label_value).strip(),
                ),
            )
        )
    if not records:
        raise ValueError(f"No valid protein records were found in {xlsx_path}.")
    return records


def _load_esm_components(model_name: str, device: Optional[str] = None):
    """Load tokenizer, model, and runtime device lazily."""
    try:
        import torch
        from transformers import AutoTokenizer, EsmModel
    except ImportError as exc:
        raise RuntimeError(
            "ESM analysis requires optional dependencies. "
            "Install them with `pip install 'ariadne-tps[esm]'` or "
            "`pip install torch transformers`."
        ) from exc

    runtime_device = device or ("cuda" if torch.cuda.is_available() else "cpu")
    tokenizer = AutoTokenizer.from_pretrained(model_name)
    model = EsmModel.from_pretrained(model_name)
    model.eval()
    model.to(runtime_device)
    return torch, tokenizer, model, runtime_device


def _load_torch():
    """Load torch lazily for the MLP classifier."""
    try:
        import torch
    except ImportError as exc:
        raise RuntimeError(
            "The trainable CeeSs classifier requires torch. "
            "Install it with `pip install 'ariadne-tps[esm]'` or `pip install torch`."
        ) from exc
    return torch


def _resolve_torch_device(torch_module, device: Optional[str] = None) -> str:
    """Resolve the runtime torch device string safely."""
    if device:
        if str(device).startswith("cuda") and not torch_module.cuda.is_available():
            logger.warning("Requested torch device %s but CUDA is unavailable; falling back to cpu.", device)
            return "cpu"
        return str(device)
    return "cuda" if torch_module.cuda.is_available() else "cpu"


def compute_esm_embeddings(
    records: list[TPSTypeRecord],
    *,
    model_name: str = DEFAULT_ESM_MODEL_NAME,
    batch_size: int = 4,
    max_length: int = 2048,
    device: Optional[str] = None,
) -> np.ndarray:
    """Compute mean-pooled ESM2 embeddings for a set of protein sequences."""
    model_name = resolve_esm_model_name(model_name)
    torch, tokenizer, model, runtime_device = _load_esm_components(model_name, device=device)
    all_embeddings: list[np.ndarray] = []

    for start in range(0, len(records), batch_size):
        batch = records[start : start + batch_size]
        sequences = [record.sequence for record in batch]
        encoded = tokenizer(
            sequences,
            return_tensors="pt",
            padding=True,
            truncation=True,
            max_length=max_length,
        )
        encoded = {key: value.to(runtime_device) for key, value in encoded.items()}
        with torch.no_grad():
            outputs = model(**encoded)
        hidden = outputs.last_hidden_state
        attention_mask = encoded["attention_mask"]

        for row_index in range(hidden.shape[0]):
            token_count = int(attention_mask[row_index].sum().item())
            if token_count <= 2:
                pooled = hidden[row_index, :token_count].mean(dim=0)
            else:
                # Exclude BOS/EOS tokens when computing the sequence embedding.
                pooled = hidden[row_index, 1 : token_count - 1].mean(dim=0)
            all_embeddings.append(pooled.detach().cpu().numpy())

    return np.vstack(all_embeddings)


class _TorchMLPClassifier(BaseEstimator, ClassifierMixin):
    """Small MLP classifier trained on frozen ESM embeddings."""

    def __init__(
        self,
        *,
        epochs: int = 200,
        hidden_dim: int = 128,
        dropout: float = 0.1,
        learning_rate: float = 1e-3,
        weight_decay: float = 1e-4,
        batch_size: int = 8,
        random_state: int = 0,
        device: Optional[str] = None,
    ):
        self.epochs = epochs
        self.hidden_dim = hidden_dim
        self.dropout = dropout
        self.learning_rate = learning_rate
        self.weight_decay = weight_decay
        self.batch_size = batch_size
        self.random_state = random_state
        self.device = device

    def _build_model(self, torch_module, input_dim: int, output_dim: int):
        return torch_module.nn.Sequential(
            torch_module.nn.Linear(input_dim, self.hidden_dim),
            torch_module.nn.ReLU(),
            torch_module.nn.Dropout(self.dropout),
            torch_module.nn.Linear(self.hidden_dim, output_dim),
        )

    def _transform(self, embeddings: np.ndarray) -> np.ndarray:
        matrix = np.asarray(embeddings, dtype=np.float32)
        return (matrix - self.mean_) / self.scale_

    def fit(self, embeddings: np.ndarray, labels: list[str], *, progress_bar=None, progress_label: Optional[str] = None):
        torch_module = _load_torch()
        device = _resolve_torch_device(torch_module, self.device)
        x = np.asarray(embeddings, dtype=np.float32)
        y = np.asarray(labels)
        if x.ndim != 2:
            raise ValueError("Embeddings for the MLP classifier must be a 2D matrix.")
        self.classes_ = np.array(sorted({str(label) for label in y}))
        if len(self.classes_) < 2:
            raise ValueError("The MLP classifier requires at least two classes.")
        class_to_index = {label: index for index, label in enumerate(self.classes_)}
        y_index = np.array([class_to_index[str(label)] for label in y], dtype=np.int64)

        self.mean_ = x.mean(axis=0, keepdims=True)
        self.scale_ = x.std(axis=0, keepdims=True)
        self.scale_[self.scale_ == 0] = 1.0
        x_scaled = self._transform(x)

        counts = np.bincount(y_index, minlength=len(self.classes_)).astype(np.float32)
        counts[counts == 0] = 1.0
        weights = len(y_index) / (len(self.classes_) * counts)

        torch_module.manual_seed(int(self.random_state))
        if torch_module.cuda.is_available():
            torch_module.cuda.manual_seed_all(int(self.random_state))

        x_tensor = torch_module.tensor(x_scaled, dtype=torch_module.float32)
        y_tensor = torch_module.tensor(y_index, dtype=torch_module.long)
        dataset = torch_module.utils.data.TensorDataset(x_tensor, y_tensor)
        loader = torch_module.utils.data.DataLoader(
            dataset,
            batch_size=max(1, min(int(self.batch_size), len(dataset))),
            shuffle=True,
        )

        self.model_ = self._build_model(torch_module, x.shape[1], len(self.classes_)).to(device)
        criterion = torch_module.nn.CrossEntropyLoss(
            weight=torch_module.tensor(weights, dtype=torch_module.float32, device=device)
        )
        optimizer = torch_module.optim.AdamW(
            self.model_.parameters(),
            lr=float(self.learning_rate),
            weight_decay=float(self.weight_decay),
        )

        self.device_ = device
        self.training_epochs_ = max(1, int(self.epochs))
        owns_progress = progress_bar is None
        epoch_iterator = (
            tqdm(
                range(self.training_epochs_),
                desc=progress_label or "Stage 3: MLP classifier",
                unit="epoch",
                leave=True,
                dynamic_ncols=True,
            )
            if owns_progress
            else range(self.training_epochs_)
        )
        for epoch_index in epoch_iterator:
            self.model_.train()
            running_loss = 0.0
            step_count = 0
            for batch_x, batch_y in loader:
                batch_x = batch_x.to(device)
                batch_y = batch_y.to(device)
                optimizer.zero_grad()
                logits = self.model_(batch_x)
                loss = criterion(logits, batch_y)
                loss.backward()
                optimizer.step()
                running_loss += float(loss.item())
                step_count += 1
            latest_loss = running_loss / max(step_count, 1)
            if owns_progress and hasattr(epoch_iterator, "set_postfix"):
                epoch_iterator.set_postfix(loss=f"{latest_loss:.4f}")
            if progress_bar is not None:
                if hasattr(progress_bar, "set_description_str"):
                    progress_bar.set_description_str(progress_label or "Stage 3: MLP classifier")
                if hasattr(progress_bar, "set_postfix"):
                    progress_bar.set_postfix(loss=f"{latest_loss:.4f}")
                progress_bar.update(1)
        return self

    def predict_proba(self, embeddings: np.ndarray) -> np.ndarray:
        torch_module = _load_torch()
        x_scaled = self._transform(embeddings)
        x_tensor = torch_module.tensor(x_scaled, dtype=torch_module.float32, device=self.device_)
        self.model_.eval()
        with torch_module.no_grad():
            logits = self.model_(x_tensor)
            probabilities = torch_module.softmax(logits, dim=1)
        return probabilities.detach().cpu().numpy()

    def decision_function(self, embeddings: np.ndarray) -> np.ndarray:
        torch_module = _load_torch()
        x_scaled = self._transform(embeddings)
        x_tensor = torch_module.tensor(x_scaled, dtype=torch_module.float32, device=self.device_)
        self.model_.eval()
        with torch_module.no_grad():
            logits = self.model_(x_tensor)
        return logits.detach().cpu().numpy()

    def predict(self, embeddings: np.ndarray) -> np.ndarray:
        probabilities = self.predict_proba(embeddings)
        indices = probabilities.argmax(axis=1)
        return self.classes_[indices]


def _save_mlp_classifier_checkpoint(path: PathLike, fitted: _TorchMLPClassifier) -> Path:
    """Persist a fitted Torch MLP classifier so later runs can reuse it."""
    torch = _load_torch()
    target = Path(path)
    state_dict = {key: value.detach().cpu() for key, value in fitted.model_.state_dict().items()}
    torch.save(
        {
            "classes": [str(value) for value in fitted.classes_],
            "mean": np.asarray(fitted.mean_, dtype=np.float32),
            "scale": np.asarray(fitted.scale_, dtype=np.float32),
            "hidden_dim": int(fitted.hidden_dim),
            "dropout": float(fitted.dropout),
            "training_epochs": int(getattr(fitted, "training_epochs_", fitted.epochs)),
            "learning_rate": float(fitted.learning_rate),
            "weight_decay": float(fitted.weight_decay),
            "batch_size": int(fitted.batch_size),
            "state_dict": state_dict,
        },
        target,
    )
    return target


def _load_mlp_classifier_checkpoint(path: PathLike, *, device: Optional[str] = None) -> _TorchMLPClassifier:
    """Load a persisted Torch MLP classifier checkpoint."""
    torch = _load_torch()
    checkpoint = torch.load(Path(path), map_location="cpu", weights_only=False)
    fitted = _TorchMLPClassifier(
        epochs=int(checkpoint.get("training_epochs", 1)),
        hidden_dim=int(checkpoint["hidden_dim"]),
        dropout=float(checkpoint["dropout"]),
        learning_rate=float(checkpoint.get("learning_rate", 1e-3)),
        weight_decay=float(checkpoint.get("weight_decay", 1e-4)),
        batch_size=int(checkpoint.get("batch_size", 8)),
        random_state=0,
        device=device,
    )
    fitted.classes_ = np.array([str(value) for value in checkpoint["classes"]], dtype=object)
    fitted.mean_ = np.asarray(checkpoint["mean"], dtype=np.float32)
    fitted.scale_ = np.asarray(checkpoint["scale"], dtype=np.float32)
    runtime_device = _resolve_torch_device(torch, device)
    input_dim = int(fitted.mean_.shape[1])
    fitted.model_ = fitted._build_model(torch, input_dim, len(fitted.classes_)).to(runtime_device)
    fitted.model_.load_state_dict(checkpoint["state_dict"])
    fitted.model_.eval()
    fitted.device_ = runtime_device
    fitted.training_epochs_ = int(checkpoint.get("training_epochs", 1))
    return fitted


def _build_classifier(
    *,
    classifier_kind: str = "mlp",
    random_state: int = 0,
    epochs: int = 200,
    hidden_dim: int = 128,
    dropout: float = 0.1,
    learning_rate: float = 1e-3,
    weight_decay: float = 1e-4,
    batch_size: int = 8,
    device: Optional[str] = None,
):
    """Construct the requested small-sample type classifier."""
    if classifier_kind == "logreg":
        return Pipeline(
            steps=[
                ("scaler", StandardScaler()),
                (
                    "classifier",
                    LogisticRegression(
                        max_iter=4000,
                        class_weight="balanced",
                        random_state=random_state,
                    ),
                ),
            ]
        )
    if classifier_kind == "mlp":
        return _TorchMLPClassifier(
            epochs=epochs,
            hidden_dim=hidden_dim,
            dropout=dropout,
            learning_rate=learning_rate,
            weight_decay=weight_decay,
            batch_size=batch_size,
            random_state=random_state,
            device=device,
        )
    raise ValueError(f"Unsupported classifier kind: {classifier_kind}")


def _projection_coordinates(embeddings: np.ndarray, labels: list[str]) -> tuple[np.ndarray, str]:
    """Compute a 2D visualization projection, preferring supervised LDA."""
    unique_labels = sorted(set(labels))
    if len(unique_labels) >= 2:
        max_components = min(2, len(unique_labels) - 1)
        if max_components >= 1:
            try:
                lda = LinearDiscriminantAnalysis(n_components=max_components)
                coords = lda.fit_transform(embeddings, labels)
                if coords.shape[1] == 1:
                    coords = np.pad(coords, ((0, 0), (0, 1)))
                return coords[:, :2], "lda"
            except Exception as error:
                logger.warning("LDA projection failed; falling back to PCA: %s", error)
    coords = PCA(n_components=2, random_state=0).fit_transform(embeddings)
    return coords, "pca"


def _fit_projection_model(embeddings: np.ndarray, labels: list[str]) -> tuple[object, str]:
    """Fit a reusable 2D projection model for training and query embeddings."""
    unique_labels = sorted(set(labels))
    if len(unique_labels) >= 2:
        max_components = min(2, len(unique_labels) - 1)
        if max_components >= 1:
            try:
                lda = LinearDiscriminantAnalysis(n_components=max_components)
                lda.fit(embeddings, labels)
                return lda, "lda"
            except Exception as error:
                logger.warning("LDA projection fit failed; falling back to PCA: %s", error)
    pca = PCA(n_components=2, random_state=0)
    pca.fit(embeddings)
    return pca, "pca"


def _transform_projection(model: object, embeddings: np.ndarray) -> np.ndarray:
    """Project embeddings with a fitted LDA/PCA model."""
    coords = np.asarray(model.transform(embeddings))
    if coords.ndim == 1:
        coords = coords[:, None]
    if coords.shape[1] == 1:
        coords = np.pad(coords, ((0, 0), (0, 1)))
    return coords[:, :2]


def _cross_validated_metrics(
    embeddings: np.ndarray,
    labels: list[str],
    *,
    cv_folds: int,
    random_state: int,
    classifier_kind: str = "mlp",
    epochs: int = 200,
    hidden_dim: int = 128,
    dropout: float = 0.1,
    learning_rate: float = 1e-3,
    weight_decay: float = 1e-4,
    batch_size: int = 8,
    device: Optional[str] = None,
) -> tuple[np.ndarray, np.ndarray, list[dict[str, object]], list[dict[str, object]], int]:
    """Compute cross-validated predictions plus metric and confusion tables."""
    class_counts = Counter(labels)
    min_class_size = min(class_counts.values())
    n_splits = max(2, min(cv_folds, min_class_size))
    classifier = _build_classifier(
        classifier_kind=classifier_kind,
        random_state=random_state,
        epochs=epochs,
        hidden_dim=hidden_dim,
        dropout=dropout,
        learning_rate=learning_rate,
        weight_decay=weight_decay,
        batch_size=batch_size,
        device=device,
    )
    splitter = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=random_state)
    cv_predictions = cross_val_predict(classifier, embeddings, labels, cv=splitter, method="predict")
    cv_probabilities = cross_val_predict(classifier, embeddings, labels, cv=splitter, method="predict_proba")

    label_order = sorted(class_counts)
    confusion = confusion_matrix(labels, cv_predictions, labels=label_order)
    confusion_rows = []
    for row_label, row_values in zip(label_order, confusion):
        row = {"true_type": row_label}
        for column_label, value in zip(label_order, row_values):
            row[column_label] = int(value)
        confusion_rows.append(row)

    report = classification_report(labels, cv_predictions, labels=label_order, output_dict=True, zero_division=0)
    metrics_rows = [
        {"metric": "num_sequences", "value": len(labels)},
        {"metric": "num_types", "value": len(class_counts)},
        {"metric": "cv_folds", "value": n_splits},
        {"metric": "classifier_kind", "value": classifier_kind},
        {"metric": "cv_accuracy", "value": round(float(accuracy_score(labels, cv_predictions)), 6)},
        {"metric": "cv_macro_f1", "value": round(float(f1_score(labels, cv_predictions, average="macro")), 6)},
        {"metric": "cv_weighted_f1", "value": round(float(f1_score(labels, cv_predictions, average="weighted")), 6)},
    ]
    if classifier_kind == "mlp":
        metrics_rows.extend(
            [
                {"metric": "mlp_epochs", "value": int(epochs)},
                {"metric": "mlp_hidden_dim", "value": int(hidden_dim)},
                {"metric": "mlp_dropout", "value": round(float(dropout), 6)},
                {"metric": "mlp_learning_rate", "value": round(float(learning_rate), 8)},
                {"metric": "mlp_weight_decay", "value": round(float(weight_decay), 8)},
                {"metric": "mlp_batch_size", "value": int(batch_size)},
            ]
        )
    for label in label_order:
        label_report = report.get(label, {})
        metrics_rows.extend(
            [
                {"metric": f"{label}_precision", "value": round(float(label_report.get("precision", 0.0)), 6)},
                {"metric": f"{label}_recall", "value": round(float(label_report.get("recall", 0.0)), 6)},
                {"metric": f"{label}_f1", "value": round(float(label_report.get("f1-score", 0.0)), 6)},
                {"metric": f"{label}_support", "value": int(label_report.get("support", 0))},
            ]
        )
    return cv_predictions, cv_probabilities, metrics_rows, confusion_rows, n_splits


def _color_map(labels: list[str]) -> dict[str, str]:
    """Assign stable colors to type labels."""
    palette = [
        "#2563EB",
        "#0F766E",
        "#C2410C",
        "#7C3AED",
        "#DC2626",
        "#0891B2",
        "#65A30D",
        "#BE185D",
    ]
    return {label: palette[index % len(palette)] for index, label in enumerate(sorted(set(labels)))}


def _classifier_classes(fitted) -> list[str]:
    """Return class labels from either sklearn pipeline or torch estimator."""
    if hasattr(fitted, "named_steps"):
        return list(fitted.named_steps["classifier"].classes_)
    return list(fitted.classes_)


def _type_probability_column(label: str) -> str:
    """Return a stable TSV column name for one TPS-type probability."""
    slug = re.sub(r"[^a-z0-9]+", "_", str(label).strip().lower()).strip("_")
    return f"esm_type_probability_{slug or 'unknown'}"


def _type_rawscore_column(label: str) -> str:
    """Return a stable TSV column name for one TPS-type raw score."""
    slug = re.sub(r"[^a-z0-9]+", "_", str(label).strip().lower()).strip("_")
    return f"esm_type_rawscore_{slug or 'unknown'}"


def _type_output_stem(label: str) -> str:
    """Return a stable filesystem-friendly stem for one TPS type."""
    return re.sub(r"[^a-z0-9]+", "_", str(label).strip().lower()).strip("_") or "unknown"


def _classifier_raw_scores(fitted, embeddings: np.ndarray) -> np.ndarray:
    """Return pre-softmax / pre-normalization scores from either classifier family."""
    if hasattr(fitted, "decision_function"):
        raw = fitted.decision_function(embeddings)
    elif hasattr(fitted, "named_steps") and hasattr(fitted.named_steps.get("classifier"), "decision_function"):
        raw = fitted.named_steps["classifier"].decision_function(fitted.named_steps["scaler"].transform(embeddings))
    else:
        raise ValueError("The fitted classifier does not expose raw scores.")
    matrix = np.asarray(raw, dtype=np.float32)
    if matrix.ndim == 1:
        classes = _classifier_classes(fitted)
        if len(classes) == 2:
            matrix = np.column_stack([-matrix, matrix])
        else:
            matrix = matrix[:, None]
    return matrix


def _collapse_labels_to_ceess_group(labels: list[str] | np.ndarray, positive_labels: set[str]) -> list[str]:
    """Collapse detailed type labels into the CeeSs / non-CeeSs grouping."""
    return ["CeeSs" if str(label) in positive_labels else "non-CeeSs" for label in labels]


def _render_type_svg(
    records: list[TPSTypeRecord],
    labels: list[str],
    coords: np.ndarray,
    output_path: PathLike,
    *,
    projection_method: str,
) -> Path:
    """Render a clean 2D SVG showing ESM type separation."""
    target = Path(output_path)
    width = 1100
    height = 760
    left = 92
    top = 74
    right = 280
    bottom = 104
    plot_width = width - left - right
    plot_height = height - top - bottom
    plot_right = left + plot_width
    plot_bottom = top + plot_height

    xs = coords[:, 0]
    ys = coords[:, 1]
    min_x, max_x = float(xs.min()), float(xs.max())
    min_y, max_y = float(ys.min()), float(ys.max())
    pad_x = (max_x - min_x) * 0.08 or 0.5
    pad_y = (max_y - min_y) * 0.08 or 0.5
    min_x -= pad_x
    max_x += pad_x
    min_y -= pad_y
    max_y += pad_y

    def scale_x(value: float) -> float:
        return left + ((value - min_x) / max(max_x - min_x, 1e-9)) * plot_width

    def scale_y(value: float) -> float:
        return plot_bottom - ((value - min_y) / max(max_y - min_y, 1e-9)) * plot_height

    colors = _color_map(labels)
    svg: list[str] = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}" font-family="Helvetica, Arial, sans-serif">',
        f'<rect width="{width}" height="{height}" fill="white" />',
        f'<text x="{left}" y="38" font-size="28" font-weight="700" fill="#0B132B">Ariadne ESM Type Embedding</text>',
        f'<text x="{left}" y="60" font-size="12" fill="#64748B">Coral TPS proteins from TPS.xlsx projected with {projection_method.upper()} on top of mean-pooled ESM2 embeddings.</text>',
        f'<rect x="{left}" y="{top}" width="{plot_width}" height="{plot_height}" fill="#FFFFFF" stroke="#E2E8F0" stroke-width="1.2" />',
    ]

    for record, label, x_value, y_value in zip(records, labels, xs, ys):
        svg.append(
            f'<circle cx="{scale_x(float(x_value)):.2f}" cy="{scale_y(float(y_value)):.2f}" r="6.8" '
            f'fill="{colors[label]}" fill-opacity="0.9" stroke="white" stroke-width="1.2">'
            f'<title>{record.name} | {label}</title></circle>'
        )

    legend_x = plot_right + 34
    legend_y = top + 18
    svg.append(f'<text x="{legend_x}" y="{legend_y}" font-size="13" font-weight="700" fill="#0F172A">TPS types</text>')
    current_y = legend_y + 24
    for label in sorted(set(labels)):
        svg.append(f'<circle cx="{legend_x + 7}" cy="{current_y - 4}" r="6" fill="{colors[label]}" />')
        svg.append(f'<text x="{legend_x + 20}" y="{current_y}" font-size="12" fill="#334155">{label}</text>')
        current_y += 24

    svg.append(
        f'<text x="{left}" y="{height - 28}" font-size="10" fill="#94A3B8" font-style="italic">'
        f'Point labels are hidden to keep the figure readable. Inspect esm_projection.tsv and esm_predictions.tsv for per-record details.</text>'
    )
    svg.append("</svg>")
    target.write_text("".join(svg))
    return target


def _render_ceess_candidate_svg(
    train_records: list[TPSTypeRecord],
    train_labels: list[str],
    train_coords: np.ndarray,
    candidate_rows: list[dict[str, object]],
    candidate_coords: np.ndarray,
    output_path: PathLike,
    *,
    projection_method: str,
) -> Path:
    """Render a combined SVG with labeled TPS references and coral-like candidates."""
    target = Path(output_path)
    width = 1180
    height = 820
    left = 92
    top = 74
    right = 320
    bottom = 108
    plot_width = width - left - right
    plot_height = height - top - bottom
    plot_right = left + plot_width
    plot_bottom = top + plot_height

    all_coords = np.vstack([train_coords, candidate_coords]) if len(candidate_rows) else train_coords
    xs = all_coords[:, 0]
    ys = all_coords[:, 1]
    min_x, max_x = float(xs.min()), float(xs.max())
    min_y, max_y = float(ys.min()), float(ys.max())
    pad_x = (max_x - min_x) * 0.08 or 0.5
    pad_y = (max_y - min_y) * 0.08 or 0.5
    min_x -= pad_x
    max_x += pad_x
    min_y -= pad_y
    max_y += pad_y

    def scale_x(value: float) -> float:
        return left + ((value - min_x) / max(max_x - min_x, 1e-9)) * plot_width

    def scale_y(value: float) -> float:
        return plot_bottom - ((value - min_y) / max(max_y - min_y, 1e-9)) * plot_height

    colors = _color_map(train_labels)
    svg: list[str] = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}" font-family="Helvetica, Arial, sans-serif">',
        f'<rect width="{width}" height="{height}" fill="white" />',
        f'<text x="{left}" y="38" font-size="28" font-weight="700" fill="#0B132B">Ariadne CeeSs Candidate Embedding</text>',
        f'<text x="{left}" y="60" font-size="12" fill="#64748B">Coral TPS references from TPS.xlsx plus coral-like candidates projected with {projection_method.upper()} on mean-pooled ESM2 embeddings.</text>',
        f'<rect x="{left}" y="{top}" width="{plot_width}" height="{plot_height}" fill="#FFFFFF" stroke="#E2E8F0" stroke-width="1.2" />',
    ]

    for record, label, x_value, y_value in zip(train_records, train_labels, train_coords[:, 0], train_coords[:, 1]):
        svg.append(
            f'<circle cx="{scale_x(float(x_value)):.2f}" cy="{scale_y(float(y_value)):.2f}" r="5.8" '
            f'fill="{colors[label]}" fill-opacity="0.48" stroke="white" stroke-width="1.0">'
            f'<title>{record.name} | reference | {label}</title></circle>'
        )

    for row, x_value, y_value in zip(candidate_rows, candidate_coords[:, 0], candidate_coords[:, 1]):
        label = str(row.get("esm_ceess_label", row.get("esm_ceess_subtype", "")))
        fill = colors.get(label, "#111827")
        x = scale_x(float(x_value))
        y = scale_y(float(y_value))
        points = [
            f"{x:.2f},{y - 8.0:.2f}",
            f"{x + 8.0:.2f},{y:.2f}",
            f"{x:.2f},{y + 8.0:.2f}",
            f"{x - 8.0:.2f},{y:.2f}",
        ]
        stroke = "#111827" if row["is_ceess_candidate"] == "yes" else "#94A3B8"
        svg.append(
            f'<polygon points="{" ".join(points)}" fill="{fill}" fill-opacity="0.92" stroke="{stroke}" stroke-width="1.4">'
            f'<title>{row["sequence_id"]} | candidate | top={row["esm_type_prediction"]} | CeeSs={row["esm_ceess_probability"]}</title></polygon>'
        )

    legend_x = plot_right + 34
    legend_y = top + 18
    svg.append(f'<text x="{legend_x}" y="{legend_y}" font-size="13" font-weight="700" fill="#0F172A">Reference TPS types</text>')
    current_y = legend_y + 24
    for label in sorted(set(train_labels)):
        svg.append(f'<circle cx="{legend_x + 7}" cy="{current_y - 4}" r="6" fill="{colors[label]}" fill-opacity="0.68" />')
        svg.append(f'<text x="{legend_x + 20}" y="{current_y}" font-size="12" fill="#334155">{label}</text>')
        current_y += 24

    current_y += 14
    svg.append(f'<text x="{legend_x}" y="{current_y}" font-size="13" font-weight="700" fill="#0F172A">Coral-like candidates</text>')
    current_y += 24
    diamond = [
        f"{legend_x + 7:.2f},{current_y - 12:.2f}",
        f"{legend_x + 15:.2f},{current_y - 4:.2f}",
        f"{legend_x + 7:.2f},{current_y + 4:.2f}",
        f"{legend_x - 1:.2f},{current_y - 4:.2f}",
    ]
    svg.append(f'<polygon points="{" ".join(diamond)}" fill="#64748B" stroke="#111827" stroke-width="1.2" />')
    svg.append(f'<text x="{legend_x + 24}" y="{current_y}" font-size="12" fill="#334155">ESM-scored coral-like candidates</text>')
    current_y += 24
    svg.append(f'<text x="{legend_x}" y="{current_y}" font-size="11" fill="#64748B">Dark outline = final CeeSs candidate</text>')

    svg.append(
        f'<text x="{left}" y="{height - 28}" font-size="10" fill="#94A3B8" font-style="italic">'
        f'Candidate labels are hidden to keep the figure readable. Inspect ceess_predictions.tsv and ceess_candidates.tsv for per-record details.</text>'
    )
    svg.append("</svg>")
    target.write_text("".join(svg))
    return target


def analyze_tps_types_with_esm(
    xlsx_path: PathLike,
    output_dir: PathLike,
    *,
    sheet_name: Optional[str] = None,
    model_name: str = DEFAULT_ESM_MODEL_NAME,
    batch_size: int = 4,
    max_length: int = 2048,
    device: Optional[str] = None,
    cv_folds: int = 5,
    random_state: int = 0,
    classifier_kind: str = "mlp",
    epochs: int = 200,
    hidden_dim: int = 128,
    dropout: float = 0.1,
    learning_rate: float = 1e-3,
    weight_decay: float = 1e-4,
    train_batch_size: int = 8,
) -> dict[str, Path]:
    """Run a supervised ESM embedding workflow for coral TPS type separation."""
    model_name = resolve_esm_model_name(model_name)
    records = load_tps_xlsx(xlsx_path, sheet_name=sheet_name)
    labels = [record.label for record in records]
    class_counts = Counter(labels)
    logger.info("Loaded %d TPS proteins across %d types from %s", len(records), len(class_counts), xlsx_path)

    embeddings = compute_esm_embeddings(
        records,
        model_name=model_name,
        batch_size=batch_size,
        max_length=max_length,
        device=device,
    )

    cv_predictions, cv_probabilities, metrics_rows, confusion_rows, n_splits = _cross_validated_metrics(
        embeddings,
        labels,
        cv_folds=cv_folds,
        random_state=random_state,
        classifier_kind=classifier_kind,
        epochs=epochs,
        hidden_dim=hidden_dim,
        dropout=dropout,
        learning_rate=learning_rate,
        weight_decay=weight_decay,
        batch_size=train_batch_size,
        device=device,
    )

    classifier = _build_classifier(
        classifier_kind=classifier_kind,
        random_state=random_state,
        epochs=epochs,
        hidden_dim=hidden_dim,
        dropout=dropout,
        learning_rate=learning_rate,
        weight_decay=weight_decay,
        batch_size=train_batch_size,
        device=device,
    )
    fitted = clone(classifier).fit(embeddings, labels)
    full_predictions = fitted.predict(embeddings)
    full_probabilities = fitted.predict_proba(embeddings)
    classes = _classifier_classes(fitted)

    coords, projection_method = _projection_coordinates(embeddings, labels)
    output_root = ensure_directory(output_dir)
    np.savez_compressed(
        output_root / "esm_embeddings.npz",
        embeddings=embeddings,
        labels=np.array(labels),
        names=np.array([record.name for record in records]),
    )

    projection_rows = []
    prediction_rows = []
    for index, record in enumerate(records):
        projection_rows.append(
            {
                "name": record.name,
                "type": record.label,
                "species": record.species,
                "length": len(record.sequence),
                "projection_method": projection_method,
                "axis1": round(float(coords[index, 0]), 6),
                "axis2": round(float(coords[index, 1]), 6),
            }
        )
        cv_pred = str(cv_predictions[index])
        full_pred = str(full_predictions[index])
        prediction_rows.append(
            {
                "name": record.name,
                "true_type": record.label,
                "cv_predicted_type": cv_pred,
                "cv_confidence": round(float(cv_probabilities[index].max()), 6),
                "full_predicted_type": full_pred,
                "full_confidence": round(float(full_probabilities[index].max()), 6),
                "species": record.species,
                "length": len(record.sequence),
            }
        )

    label_order = sorted(class_counts)
    metrics_rows.append({"metric": "esm_model", "value": model_name})

    centroid_rows = []
    for label in label_order:
        indices = [index for index, current in enumerate(labels) if current == label]
        centroid = embeddings[indices].mean(axis=0)
        centroid_rows.append(
            {
                "type": label,
                "count": len(indices),
                "axis1": round(float(coords[indices, 0].mean()), 6),
                "axis2": round(float(coords[indices, 1].mean()), 6),
                "embedding_norm": round(float(np.linalg.norm(centroid)), 6),
            }
        )

    projection_path = write_tsv(projection_rows, output_root / "esm_projection.tsv")
    predictions_path = write_tsv(prediction_rows, output_root / "esm_predictions.tsv")
    confusion_path = write_tsv(confusion_rows, output_root / "esm_confusion_matrix.tsv")
    metrics_path = write_tsv(metrics_rows, output_root / "esm_metrics.tsv")
    centroids_path = write_tsv(centroid_rows, output_root / "esm_type_centroids.tsv")
    svg_path = _render_type_svg(
        records,
        labels,
        coords,
        output_root / "esm_embedding.svg",
        projection_method=projection_method,
    )

    return {
        "embeddings": output_root / "esm_embeddings.npz",
        "projection": projection_path,
        "predictions": predictions_path,
        "confusion_matrix": confusion_path,
        "metrics": metrics_path,
        "type_centroids": centroids_path,
        "embedding_svg": svg_path,
    }


def classify_ceess_candidates_with_esm(
    candidate_records: list[FastaRecord],
    xlsx_path: PathLike,
    output_dir: PathLike,
    *,
    sheet_name: Optional[str] = None,
    model_name: str = DEFAULT_ESM_MODEL_NAME,
    batch_size: int = 4,
    max_length: int = 2048,
    device: Optional[str] = None,
    cv_folds: int = 5,
    random_state: int = 0,
    ceess_threshold: float = 0.5,
    classifier_kind: str = "mlp",
    epochs: int = 200,
    hidden_dim: int = 128,
    representation_dim: Optional[int] = None,
    projection_dim: Optional[int] = None,
    dropout: float = 0.1,
    learning_rate: float = 1e-3,
    weight_decay: float = 1e-4,
    train_batch_size: int = 8,
    barlow_redundancy_weight: float = 0.005,
    classifier_checkpoint: Optional[PathLike] = None,
) -> CeessPredictionResult:
    """Score coral-like candidates with a multi-class ESM classifier trained on TPS types."""
    if classifier_kind == "contrastive":
        if classifier_checkpoint is not None:
            raise ValueError(
                "Loading a standalone MLP checkpoint is not supported with classifier_kind='contrastive'. "
                "Please rerun the contrastive pipeline or add a dedicated SupCon checkpoint loader."
            )
        return classify_ceess_candidates_with_supcon(
            candidate_records,
            xlsx_path,
            output_dir,
            sheet_name=sheet_name,
            model_name=model_name,
            batch_size=batch_size,
            max_length=max_length,
            device=device,
            cv_folds=cv_folds,
            random_state=random_state,
            ceess_threshold=ceess_threshold,
            epochs=epochs,
            hidden_dim=hidden_dim,
            representation_dim=representation_dim,
            projection_dim=projection_dim,
            dropout=dropout,
            learning_rate=learning_rate,
            weight_decay=weight_decay,
            train_batch_size=train_batch_size,
            redundancy_weight=barlow_redundancy_weight,
        )
    model_name = resolve_esm_model_name(model_name)
    if not candidate_records:
        raise ValueError("No coral-like candidate records were provided for CeeSs scoring.")

    train_records = load_tps_xlsx(xlsx_path, sheet_name=sheet_name)
    train_labels = [record.label for record in train_records]
    ceess_positive_labels = sorted({record.label for record in train_records if record.ceess_group == "CeeSs"})
    if not ceess_positive_labels:
        ceess_positive_labels = sorted(CEESS_POSITIVE_LABELS.intersection(train_labels))
    if not ceess_positive_labels:
        raise ValueError(
            "No CeeSs-positive labels were found in TPS.xlsx. "
            "Add a CeeSs_group/CeeSs column or include known positive labels."
        )
    query_type_records = [
        TPSTypeRecord(
            name=record.id,
            sequence=record.sequence,
            label="candidate",
            species=record.metadata.get("source", "candidate"),
            ceess_group="candidate",
        )
        for record in candidate_records
    ]
    combined_records = train_records + query_type_records
    combined_embeddings = compute_esm_embeddings(
        combined_records,
        model_name=model_name,
        batch_size=batch_size,
        max_length=max_length,
        device=device,
    )
    train_embeddings = combined_embeddings[: len(train_records)]
    query_embeddings = combined_embeddings[len(train_records) :]

    cv_predictions, _, metrics_rows, confusion_rows, _ = _cross_validated_metrics(
        train_embeddings,
        train_labels,
        cv_folds=cv_folds,
        random_state=random_state,
        classifier_kind=classifier_kind,
        epochs=epochs,
        hidden_dim=hidden_dim,
        dropout=dropout,
        learning_rate=learning_rate,
        weight_decay=weight_decay,
        batch_size=train_batch_size,
        device=device,
    )
    true_ceess_groups = _collapse_labels_to_ceess_group(train_labels, set(ceess_positive_labels))
    predicted_ceess_groups = _collapse_labels_to_ceess_group(cv_predictions, set(ceess_positive_labels))
    group_report = classification_report(
        true_ceess_groups,
        predicted_ceess_groups,
        labels=["CeeSs", "non-CeeSs"],
        output_dict=True,
        zero_division=0,
    )
    group_confusion = confusion_matrix(
        true_ceess_groups,
        predicted_ceess_groups,
        labels=["CeeSs", "non-CeeSs"],
    )
    metrics_rows.extend(
        [
            {"metric": "esm_model", "value": model_name},
            {"metric": "ceess_threshold", "value": round(float(ceess_threshold), 6)},
            {"metric": "ceess_positive_types", "value": ",".join(ceess_positive_labels)},
            {"metric": "num_ceess_references", "value": sum(record.ceess_group == "CeeSs" for record in train_records)},
            {"metric": "num_non_ceess_references", "value": sum(record.ceess_group != "CeeSs" for record in train_records)},
            {"metric": "num_coral_like_candidates", "value": len(candidate_records)},
            {"metric": "cv_ceess_group_accuracy", "value": round(float(accuracy_score(true_ceess_groups, predicted_ceess_groups)), 6)},
            {"metric": "cv_ceess_group_macro_f1", "value": round(float(f1_score(true_ceess_groups, predicted_ceess_groups, average="macro")), 6)},
            {"metric": "CeeSs_group_precision", "value": round(float(group_report.get("CeeSs", {}).get("precision", 0.0)), 6)},
            {"metric": "CeeSs_group_recall", "value": round(float(group_report.get("CeeSs", {}).get("recall", 0.0)), 6)},
            {"metric": "CeeSs_group_f1", "value": round(float(group_report.get("CeeSs", {}).get("f1-score", 0.0)), 6)},
            {"metric": "CeeSs_group_support", "value": int(group_report.get("CeeSs", {}).get("support", 0))},
            {"metric": "non_CeeSs_group_precision", "value": round(float(group_report.get("non-CeeSs", {}).get("precision", 0.0)), 6)},
            {"metric": "non_CeeSs_group_recall", "value": round(float(group_report.get("non-CeeSs", {}).get("recall", 0.0)), 6)},
            {"metric": "non_CeeSs_group_f1", "value": round(float(group_report.get("non-CeeSs", {}).get("f1-score", 0.0)), 6)},
            {"metric": "non_CeeSs_group_support", "value": int(group_report.get("non-CeeSs", {}).get("support", 0))},
        ]
    )

    output_root = ensure_directory(output_dir)
    classifier = _build_classifier(
        classifier_kind=classifier_kind,
        random_state=random_state,
        epochs=epochs,
        hidden_dim=hidden_dim,
        dropout=dropout,
        learning_rate=learning_rate,
        weight_decay=weight_decay,
        batch_size=train_batch_size,
        device=device,
    )
    classifier_checkpoint_path = Path(classifier_checkpoint) if classifier_checkpoint is not None else None
    if classifier_checkpoint_path is not None:
        if classifier_kind != "mlp":
            raise ValueError("--ceess-mlp-checkpoint currently only supports --ceess-classifier mlp.")
        fitted = _load_mlp_classifier_checkpoint(classifier_checkpoint_path, device=device)
    else:
        fitted = clone(classifier).fit(train_embeddings, train_labels)
    query_probabilities = fitted.predict_proba(query_embeddings)
    query_raw_scores = _classifier_raw_scores(fitted, query_embeddings)
    query_predictions = fitted.predict(query_embeddings)
    classes = _classifier_classes(fitted)
    class_to_index = {label: index for index, label in enumerate(classes)}

    projection_model, projection_method = _fit_projection_model(train_embeddings, train_labels)
    train_coords = _transform_projection(projection_model, train_embeddings)
    query_coords = _transform_projection(projection_model, query_embeddings)

    projection_rows = []
    prediction_rows = []
    ceess_candidate_records: list[FastaRecord] = []
    for train_record, coords in zip(train_records, train_coords):
        projection_rows.append(
            {
                "sequence_id": train_record.name,
                "role": "reference",
                "type": train_record.label,
                "species": train_record.species,
                "projection_method": projection_method,
                "axis1": round(float(coords[0]), 6),
                "axis2": round(float(coords[1]), 6),
            }
        )

    for i, (record, coords, probabilities, raw_scores, top_prediction) in enumerate(
        zip(candidate_records, query_coords, query_probabilities, query_raw_scores, query_predictions)
    ):
        ceess_probability = float(
            sum(float(probabilities[class_to_index[label]]) for label in ceess_positive_labels if label in class_to_index)
        )
        non_ceess_probability = max(0.0, 1.0 - ceess_probability)
        is_candidate = ceess_probability >= ceess_threshold
        if is_candidate:
            ceess_candidate_records.append(record.clone())
        projection_rows.append(
            {
                "sequence_id": record.id,
                "role": "candidate",
                "type": str(top_prediction),
                "species": record.metadata.get("source", "candidate"),
                "projection_method": projection_method,
                "axis1": round(float(coords[0]), 6),
                "axis2": round(float(coords[1]), 6),
            }
        )
        probability_columns = {
            _type_probability_column(label): round(float(probabilities[class_to_index[label]]), 6)
            for label in classes
            if label in class_to_index
        }
        rawscore_columns = {
            _type_rawscore_column(label): round(float(raw_scores[class_to_index[label]]), 6)
            for label in classes
            if label in class_to_index
        }
        prediction_rows.append(
            {
                "sequence_id": record.id,
                "esm_type_prediction": str(top_prediction),
                "esm_ceess_label": str(top_prediction),
                "esm_ceess_probability": round(ceess_probability, 6),
                "esm_non_ceess_probability": round(non_ceess_probability, 6),
                "is_ceess_candidate": "yes" if is_candidate else "no",
                **probability_columns,
                **rawscore_columns,
            }
        )

    prediction_rows_sorted = sorted(
        prediction_rows,
        key=lambda row: (
            0 if row["is_ceess_candidate"] == "yes" else 1,
            -float(row["esm_ceess_probability"]),
            str(row["sequence_id"]),
        )
    )

    metrics_path = write_tsv(metrics_rows, output_root / "ceess_model_metrics.tsv")
    confusion_path = write_tsv(confusion_rows, output_root / "ceess_model_confusion_matrix.tsv")
    group_confusion_path = write_tsv(
        [
            {
                "true_ceess_group": label,
                "pred_CeeSs": int(row[0]),
                "pred_non_CeeSs": int(row[1]),
            }
            for label, row in zip(["CeeSs", "non-CeeSs"], group_confusion)
        ],
        output_root / "ceess_group_confusion_matrix.tsv",
    )
    projection_path = write_tsv(projection_rows, output_root / "ceess_projection.tsv")
    predictions_path = write_tsv(prediction_rows_sorted, output_root / "ceess_predictions.tsv")
    candidates_path = write_tsv(
        [row for row in prediction_rows_sorted if row["is_ceess_candidate"] == "yes"],
        output_root / "ceess_candidates.tsv",
    )
    candidate_fasta_path = write_fasta(ceess_candidate_records, output_root / "ceess_candidates.fasta")
    subtype_tsv_dir = ensure_directory(output_root / "type_score_hits")
    subtype_fasta_dir = ensure_directory(output_root / "type_score_fastas")
    candidate_record_by_id = {record.id: record for record in candidate_records}
    for label in classes:
        probability_column = _type_probability_column(label)
        output_stem = _type_output_stem(label)
        subtype_rows = []
        subtype_fasta_records: list[FastaRecord] = []
        for row in prediction_rows_sorted:
            try:
                probability_value = float(row.get(probability_column, "") or 0.0)
            except (TypeError, ValueError):
                probability_value = 0.0
            if probability_value <= 0.95:
                continue
            subtype_rows.append(dict(row))
            record = candidate_record_by_id.get(str(row["sequence_id"]))
            if record is not None:
                subtype_fasta_records.append(record.clone())
        write_tsv(subtype_rows, subtype_tsv_dir / f"{output_stem}.tsv")
        write_fasta(subtype_fasta_records, subtype_fasta_dir / f"{output_stem}.fasta")
    checkpoint_path = None
    if classifier_kind == "mlp":
        checkpoint_path = (
            classifier_checkpoint_path
            if classifier_checkpoint_path is not None
            else _save_mlp_classifier_checkpoint(output_root / "ceess_classifier_checkpoint.pt", fitted)
        )
    embedding_path = _render_ceess_candidate_svg(
        train_records,
        train_labels,
        train_coords,
        prediction_rows,
        query_coords,
        output_root / "ceess_embedding.svg",
        projection_method=projection_method,
    )

    return CeessPredictionResult(
        output_paths={
            "ceess_model_metrics": metrics_path,
            "ceess_model_confusion_matrix": confusion_path,
            "ceess_group_confusion_matrix": group_confusion_path,
            "ceess_projection": projection_path,
            "ceess_predictions": predictions_path,
            "ceess_candidates": candidates_path,
            "ceess_candidates_fasta": candidate_fasta_path,
            "type_score_hits": subtype_tsv_dir,
            "type_score_fastas": subtype_fasta_dir,
            "ceess_embedding_svg": embedding_path,
            **({"ceess_classifier_checkpoint": checkpoint_path} if checkpoint_path is not None else {}),
        },
        prediction_rows=prediction_rows,
        candidate_ids=[row["sequence_id"] for row in prediction_rows_sorted if row["is_ceess_candidate"] == "yes"],
    )

# ---------------------------------------------------------------------------
# Barlow Twins contrastive learning variant (from ceess_supcon)
# ---------------------------------------------------------------------------

import math

class _BarlowProjectionNetwork:
    """Encoder plus projector adapted from the official Barlow Twins design."""

    def __init__(
        self,
        *,
        input_dim: int,
        hidden_dim: int,
        representation_dim: int,
        projection_dim: int,
        dropout: float,
    ):
        torch = _load_torch()
        self.torch = torch
        self.encoder = torch.nn.Sequential(
            torch.nn.Linear(input_dim, hidden_dim, bias=False),
            torch.nn.BatchNorm1d(hidden_dim),
            torch.nn.ReLU(inplace=True),
            torch.nn.Dropout(dropout),
            torch.nn.Linear(hidden_dim, representation_dim, bias=False),
            torch.nn.BatchNorm1d(representation_dim),
            torch.nn.ReLU(inplace=True),
        )
        self.projector = torch.nn.Sequential(
            torch.nn.Linear(representation_dim, hidden_dim, bias=False),
            torch.nn.BatchNorm1d(hidden_dim),
            torch.nn.ReLU(inplace=True),
            torch.nn.Dropout(dropout),
            torch.nn.Linear(hidden_dim, projection_dim, bias=False),
        )
        self.bn = torch.nn.BatchNorm1d(projection_dim, affine=False)

    def to(self, device: str):
        self.encoder.to(device)
        self.projector.to(device)
        self.bn.to(device)
        return self

    def train(self):
        self.encoder.train()
        self.projector.train()
        self.bn.train()

    def eval(self):
        self.encoder.eval()
        self.projector.eval()
        self.bn.eval()

    def parameters(self):
        return list(self.encoder.parameters()) + list(self.projector.parameters())

    def encode(self, x):
        return self.encoder(x)

    def project(self, x):
        representation = self.encode(x)
        projected = self.projector(representation)
        return representation, projected

    def state_dict(self):
        return {
            "encoder": self.encoder.state_dict(),
            "projector": self.projector.state_dict(),
            "bn": self.bn.state_dict(),
        }

    def load_state_dict(self, state_dict):
        self.encoder.load_state_dict(state_dict["encoder"])
        self.projector.load_state_dict(state_dict["projector"])
        self.bn.load_state_dict(state_dict["bn"])


def _off_diagonal(torch_module, matrix):
    size = matrix.shape[0]
    return matrix.flatten()[:-1].view(size - 1, size + 1)[:, 1:].flatten()


def _barlow_twins_loss(torch_module, model: _BarlowProjectionNetwork, z1, z2, *, redundancy_weight: float):
    """Compute the Barlow Twins redundancy-reduction objective."""
    if z1.shape != z2.shape:
        raise ValueError("Barlow Twins views must have the same shape.")
    z1 = model.bn(z1)
    z2 = model.bn(z2)
    cross_correlation = torch_module.matmul(z1.T, z2) / z1.shape[0]
    on_diag = torch_module.diagonal(cross_correlation).add_(-1.0).pow_(2).sum()
    off_diag = _off_diagonal(torch_module, cross_correlation).pow_(2).sum()
    return on_diag + float(redundancy_weight) * off_diag


def _save_supcon_checkpoint(path: Path, network: _BarlowProjectionNetwork, *, config: dict[str, object]):
    torch = _load_torch()
    state_dict = network.state_dict()
    cpu_state_dict = {}
    for module_name, module_state in state_dict.items():
        cpu_state_dict[module_name] = {key: value.detach().cpu() for key, value in module_state.items()}
    torch.save({"config": config, "state_dict": cpu_state_dict}, path)


def _save_mlp_classifier_checkpoint(path: Path, fitted: _TorchMLPClassifier):
    torch = _load_torch()
    state_dict = {key: value.detach().cpu() for key, value in fitted.model_.state_dict().items()}
    torch.save(
        {
            "classes": [str(value) for value in fitted.classes_],
            "mean": np.asarray(fitted.mean_, dtype=np.float32),
            "scale": np.asarray(fitted.scale_, dtype=np.float32),
            "hidden_dim": int(fitted.hidden_dim),
            "dropout": float(fitted.dropout),
            "training_epochs": int(fitted.training_epochs_),
            "learning_rate": float(fitted.learning_rate),
            "weight_decay": float(fitted.weight_decay),
            "batch_size": int(fitted.batch_size),
            "state_dict": state_dict,
        },
        path,
    )


class _LARS:
    """Single-process LARS optimizer adapted from the official Barlow Twins implementation."""

    def __init__(self, params, *, lr: float, weight_decay: float = 0.0, momentum: float = 0.9, eta: float = 0.001):
        torch = _load_torch()
        self.optim = torch.optim.Optimizer(
            params,
            dict(lr=lr, weight_decay=weight_decay, momentum=momentum, eta=eta),
        )

    @property
    def param_groups(self):
        return self.optim.param_groups

    def zero_grad(self):
        return self.optim.zero_grad()

    @staticmethod
    def _exclude_bias_and_norm(param) -> bool:
        return param.ndim == 1

    def step(self):
        torch = _load_torch()
        with torch.no_grad():
            for group in self.optim.param_groups:
                for param in group["params"]:
                    grad = param.grad
                    if grad is None:
                        continue
                    update = grad
                    if not self._exclude_bias_and_norm(param):
                        update = update.add(param, alpha=group["weight_decay"])
                        param_norm = torch.norm(param)
                        update_norm = torch.norm(update)
                        one = torch.ones_like(param_norm)
                        q = torch.where(
                            param_norm > 0.0,
                            torch.where(update_norm > 0.0, group["eta"] * param_norm / update_norm, one),
                            one,
                        )
                        update = update.mul(q)
                    state = self.optim.state[param]
                    if "mu" not in state:
                        state["mu"] = torch.zeros_like(param)
                    mu = state["mu"]
                    mu.mul_(group["momentum"]).add_(update)
                    param.add_(mu, alpha=-group["lr"])


def _adjust_learning_rate(
    optimizer: _LARS,
    *,
    step: int,
    steps_per_epoch: int,
    epochs: int,
    batch_size: int,
    learning_rate_weights: float,
    learning_rate_biases: float,
):
    max_steps = max(1, epochs * steps_per_epoch)
    warmup_steps = max(1, min(10 * steps_per_epoch, max_steps))
    base_lr = batch_size / 256.0
    if step < warmup_steps:
        lr = base_lr * step / warmup_steps
    else:
        shifted_step = step - warmup_steps
        shifted_max_steps = max(1, max_steps - warmup_steps)
        q = 0.5 * (1.0 + math.cos(math.pi * shifted_step / shifted_max_steps))
        end_lr = base_lr * 0.001
        lr = base_lr * q + end_lr * (1.0 - q)
    optimizer.param_groups[0]["lr"] = lr * learning_rate_weights
    optimizer.param_groups[1]["lr"] = lr * learning_rate_biases


def _fit_supcon_representation(
    embeddings: np.ndarray,
    labels: list[str],
    *,
    hidden_dim: int,
    representation_dim: Optional[int],
    projection_dim: Optional[int],
    dropout: float,
    learning_rate: float,
    weight_decay: float,
    epochs: int,
    batch_size: int,
    redundancy_weight: float,
    device: Optional[str],
    random_state: int,
    progress_bar=None,
):
    torch = _load_torch()
    runtime_device = _resolve_torch_device(torch, device)
    torch.manual_seed(int(random_state))
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(int(random_state))

    input_dim = int(embeddings.shape[1])
    if representation_dim is None:
        representation_dim = max(64, min(int(hidden_dim), 256))
    else:
        representation_dim = max(16, int(representation_dim))
    if projection_dim is None:
        projection_dim = min(128, representation_dim)
    else:
        projection_dim = max(8, min(int(projection_dim), representation_dim))
    network = _BarlowProjectionNetwork(
        input_dim=input_dim,
        hidden_dim=max(64, int(hidden_dim)),
        representation_dim=representation_dim,
        projection_dim=projection_dim,
        dropout=float(dropout),
    ).to(runtime_device)

    x_tensor = torch.tensor(np.asarray(embeddings, dtype=np.float32), dtype=torch.float32)
    dataset = torch.utils.data.TensorDataset(x_tensor)
    effective_batch_size = max(2, min(int(batch_size), len(dataset)))
    loader = torch.utils.data.DataLoader(
        dataset,
        batch_size=effective_batch_size,
        shuffle=True,
        drop_last=len(dataset) > effective_batch_size,
    )
    param_weights = []
    param_biases = []
    for param in network.parameters():
        if param.ndim == 1:
            param_biases.append(param)
        else:
            param_weights.append(param)
    optimizer = _LARS(
        [
            {"params": param_weights},
            {"params": param_biases},
        ],
        lr=0.0,
        weight_decay=float(weight_decay),
    )

    history_rows: list[dict[str, object]] = []
    owns_progress = progress_bar is None
    epoch_iterator = (
        tqdm(
            range(max(1, int(epochs))),
            desc="Stage 2: Barlow Twins",
            unit="epoch",
            leave=True,
            dynamic_ncols=True,
        )
        if owns_progress
        else range(max(1, int(epochs)))
    )
    for epoch in epoch_iterator:
        network.train()
        running_loss = 0.0
        step_count = 0
        for batch_index, (batch_x,) in enumerate(loader):
            batch_x = batch_x.to(runtime_device)
            current_step = epoch * max(len(loader), 1) + batch_index
            _adjust_learning_rate(
                optimizer,
                step=current_step,
                steps_per_epoch=max(len(loader), 1),
                epochs=max(1, int(epochs)),
                batch_size=effective_batch_size,
                learning_rate_weights=float(learning_rate),
                learning_rate_biases=float(learning_rate * 0.024),
            )
            optimizer.zero_grad()
            _, z1 = network.project(batch_x)
            _, z2 = network.project(batch_x)
            loss = _barlow_twins_loss(torch, network, z1, z2, redundancy_weight=float(redundancy_weight))
            loss.backward()
            optimizer.step()
            running_loss += float(loss.item())
            step_count += 1
        history_rows.append(
            {
                "epoch": epoch + 1,
                "barlow_loss": round(running_loss / max(step_count, 1), 8),
            }
        )
        latest_loss = history_rows[-1]["barlow_loss"]
        if owns_progress and hasattr(epoch_iterator, "set_postfix"):
            epoch_iterator.set_postfix(loss=f"{latest_loss:.4f}")
        if progress_bar is not None:
            if hasattr(progress_bar, "set_description_str"):
                progress_bar.set_description_str("Stage 2: Barlow Twins")
            if hasattr(progress_bar, "set_postfix"):
                progress_bar.set_postfix(loss=f"{latest_loss:.4f}")
            progress_bar.update(1)

    network.eval()
    with torch.no_grad():
        full_tensor = torch.tensor(np.asarray(embeddings, dtype=np.float32), dtype=torch.float32, device=runtime_device)
        representations = network.encode(full_tensor).detach().cpu().numpy()

    return network, representations, history_rows, {
        "input_dim": input_dim,
        "hidden_dim": max(64, int(hidden_dim)),
        "representation_dim": representation_dim,
        "projection_dim": projection_dim,
        "dropout": float(dropout),
        "redundancy_weight": float(redundancy_weight),
        "epochs": max(1, int(epochs)),
        "batch_size": effective_batch_size,
        "learning_rate": float(learning_rate),
        "weight_decay": float(weight_decay),
        "random_state": int(random_state),
    }


def _encode_with_barlow(network: _BarlowProjectionNetwork, embeddings: np.ndarray, *, device: Optional[str]) -> np.ndarray:
    torch = _load_torch()
    runtime_device = _resolve_torch_device(torch, device)
    network.eval()
    with torch.no_grad():
        x_tensor = torch.tensor(np.asarray(embeddings, dtype=np.float32), dtype=torch.float32, device=runtime_device)
        representations = network.encode(x_tensor).detach().cpu().numpy()
    return representations


def classify_ceess_candidates_with_supcon(
    candidate_records: list[FastaRecord],
    xlsx_path,
    output_dir,
    *,
    sheet_name: Optional[str] = None,
    model_name: str = DEFAULT_ESM_MODEL_NAME,
    batch_size: int = 4,
    max_length: int = 2048,
    device: Optional[str] = None,
    cv_folds: int = 5,
    random_state: int = 0,
    ceess_threshold: float = 0.9,
    epochs: int = 200,
    hidden_dim: int = 128,
    representation_dim: Optional[int] = None,
    projection_dim: Optional[int] = None,
    dropout: float = 0.1,
    learning_rate: float = 1e-3,
    weight_decay: float = 1e-4,
    train_batch_size: int = 8,
    redundancy_weight: float = 0.005,
) -> CeessPredictionResult:
    """Score coral-like candidates with a Barlow-Twins representation plus MLP classifier."""
    model_name = resolve_esm_model_name(model_name)
    if not candidate_records:
        raise ValueError("No coral-like candidate records were provided for CeeSs scoring.")

    output_root = ensure_directory(output_dir)
    train_records = load_tps_xlsx(xlsx_path, sheet_name=sheet_name)
    train_labels = [record.label for record in train_records]
    ceess_positive_labels = sorted({record.label for record in train_records if record.ceess_group == "CeeSs"})
    if not ceess_positive_labels:
        raise ValueError("Contrastive mode requires at least one CeeSs-positive label in TPS.xlsx.")

    query_type_records = [
        type(train_records[0])(
            name=record.id,
            sequence=record.sequence,
            label="candidate",
            species=record.metadata.get("source", "candidate"),
            ceess_group="candidate",
        )
        for record in candidate_records
    ]
    combined_records = train_records + query_type_records
    combined_embeddings = compute_esm_embeddings(
        combined_records,
        model_name=model_name,
        batch_size=batch_size,
        max_length=max_length,
        device=device,
    )
    train_embeddings = combined_embeddings[: len(train_records)]
    query_embeddings = combined_embeddings[len(train_records) :]

    np.savez_compressed(
        output_root / "ceess_embedding_cache.npz",
        train_embeddings=train_embeddings,
        query_embeddings=query_embeddings,
        train_labels=np.array(train_labels),
        train_names=np.array([record.name for record in train_records]),
        query_names=np.array([record.id for record in candidate_records]),
    )

    barlow_network, train_representations, training_history_rows, barlow_config = _fit_supcon_representation(
        train_embeddings,
        train_labels,
        hidden_dim=hidden_dim,
        representation_dim=representation_dim,
        projection_dim=projection_dim,
        dropout=dropout,
        learning_rate=learning_rate,
        weight_decay=weight_decay,
        epochs=epochs,
        batch_size=train_batch_size,
        redundancy_weight=redundancy_weight,
        device=device,
        random_state=random_state,
    )
    query_representations = _encode_with_barlow(barlow_network, query_embeddings, device=device)
    np.savez_compressed(
        output_root / "ceess_barlow_representations.npz",
        train_representations=train_representations,
        query_representations=query_representations,
        train_labels=np.array(train_labels),
        train_names=np.array([record.name for record in train_records]),
        query_names=np.array([record.id for record in candidate_records]),
    )
    write_tsv(training_history_rows, output_root / "ceess_barlow_training.tsv")
    _save_supcon_checkpoint(output_root / "ceess_barlow_checkpoint.pt", barlow_network, config=barlow_config)

    cv_predictions, _, metrics_rows, confusion_rows, _ = _cross_validated_metrics(
        train_representations,
        train_labels,
        cv_folds=cv_folds,
        random_state=random_state,
        classifier_kind="mlp",
        epochs=epochs,
        hidden_dim=hidden_dim,
        dropout=dropout,
        learning_rate=learning_rate,
        weight_decay=weight_decay,
        batch_size=train_batch_size,
        device=device,
    )
    true_ceess_groups = _collapse_labels_to_ceess_group(train_labels, set(ceess_positive_labels))
    predicted_ceess_groups = _collapse_labels_to_ceess_group(cv_predictions, set(ceess_positive_labels))
    group_report = classification_report(
        true_ceess_groups,
        predicted_ceess_groups,
        labels=["CeeSs", "non-CeeSs"],
        output_dict=True,
        zero_division=0,
    )
    group_confusion = confusion_matrix(
        true_ceess_groups,
        predicted_ceess_groups,
        labels=["CeeSs", "non-CeeSs"],
    )
    metrics_rows.extend(
        [
            {"metric": "representation_model", "value": "barlow_twins"},
            {"metric": "esm_model", "value": model_name},
            {"metric": "ceess_threshold", "value": round(float(ceess_threshold), 6)},
            {"metric": "ceess_positive_types", "value": ",".join(ceess_positive_labels)},
            {"metric": "num_ceess_references", "value": sum(record.ceess_group == "CeeSs" for record in train_records)},
            {"metric": "num_non_ceess_references", "value": sum(record.ceess_group != "CeeSs" for record in train_records)},
            {"metric": "num_coral_like_candidates", "value": len(candidate_records)},
            {"metric": "barlow_redundancy_weight", "value": barlow_config["redundancy_weight"]},
            {"metric": "barlow_representation_dim", "value": barlow_config["representation_dim"]},
            {"metric": "barlow_projection_dim", "value": barlow_config["projection_dim"]},
            {"metric": "cv_ceess_group_accuracy", "value": round(float(accuracy_score(true_ceess_groups, predicted_ceess_groups)), 6)},
            {"metric": "cv_ceess_group_macro_f1", "value": round(float(group_report.get("macro avg", {}).get("f1-score", 0.0)), 6)},
            {"metric": "CeeSs_group_precision", "value": round(float(group_report.get("CeeSs", {}).get("precision", 0.0)), 6)},
            {"metric": "CeeSs_group_recall", "value": round(float(group_report.get("CeeSs", {}).get("recall", 0.0)), 6)},
            {"metric": "CeeSs_group_f1", "value": round(float(group_report.get("CeeSs", {}).get("f1-score", 0.0)), 6)},
            {"metric": "CeeSs_group_support", "value": int(group_report.get("CeeSs", {}).get("support", 0))},
            {"metric": "non_CeeSs_group_precision", "value": round(float(group_report.get("non-CeeSs", {}).get("precision", 0.0)), 6)},
            {"metric": "non_CeeSs_group_recall", "value": round(float(group_report.get("non-CeeSs", {}).get("recall", 0.0)), 6)},
            {"metric": "non_CeeSs_group_f1", "value": round(float(group_report.get("non-CeeSs", {}).get("f1-score", 0.0)), 6)},
            {"metric": "non_CeeSs_group_support", "value": int(group_report.get("non-CeeSs", {}).get("support", 0))},
        ]
    )

    classifier = _TorchMLPClassifier(
        epochs=epochs,
        hidden_dim=hidden_dim,
        dropout=dropout,
        learning_rate=learning_rate,
        weight_decay=weight_decay,
        batch_size=train_batch_size,
        random_state=random_state,
        device=device,
    )
    fitted = classifier.fit(
        train_representations,
        train_labels,
        progress_label="Stage 3: MLP classifier",
    )
    _save_mlp_classifier_checkpoint(output_root / "ceess_classifier_checkpoint.pt", fitted)
    query_probabilities = fitted.predict_proba(query_representations)
    query_raw_scores = fitted.decision_function(query_representations)
    query_predictions = fitted.predict(query_representations)
    classes = [str(label) for label in fitted.classes_]
    class_to_index = {label: index for index, label in enumerate(classes)}

    projection_model, projection_method = _fit_projection_model(train_representations, train_labels)
    train_coords = _transform_projection(projection_model, train_representations)
    query_coords = _transform_projection(projection_model, query_representations)

    projection_rows = []
    prediction_rows = []
    ceess_candidate_records: list[FastaRecord] = []
    for train_record, coords in zip(train_records, train_coords):
        projection_rows.append(
            {
                "sequence_id": train_record.name,
                "role": "reference",
                "type": train_record.label,
                "species": train_record.species,
                "projection_method": projection_method,
                "axis1": round(float(coords[0]), 6),
                "axis2": round(float(coords[1]), 6),
            }
        )

    for record, coords, probabilities, raw_scores, top_prediction in zip(
        candidate_records,
        query_coords,
        query_probabilities,
        query_raw_scores,
        query_predictions,
    ):
        ceess_probability = float(
            sum(float(probabilities[class_to_index[label]]) for label in ceess_positive_labels if label in class_to_index)
        )
        non_ceess_probability = max(0.0, 1.0 - ceess_probability)
        is_candidate = ceess_probability >= ceess_threshold
        if is_candidate:
            ceess_candidate_records.append(record.clone())
        projection_rows.append(
            {
                "sequence_id": record.id,
                "role": "candidate",
                "type": str(top_prediction),
                "species": record.metadata.get("source", "candidate"),
                "projection_method": projection_method,
                "axis1": round(float(coords[0]), 6),
                "axis2": round(float(coords[1]), 6),
            }
        )
        probability_columns = {
            _type_probability_column(label): round(float(probabilities[class_to_index[label]]), 6)
            for label in classes
        }
        rawscore_columns = {
            _type_rawscore_column(label): round(float(raw_scores[class_to_index[label]]), 6)
            for label in classes
        }
        prediction_rows.append(
            {
                "sequence_id": record.id,
                "esm_type_prediction": str(top_prediction),
                "esm_ceess_label": str(top_prediction),
                "esm_ceess_probability": round(ceess_probability, 6),
                "esm_non_ceess_probability": round(non_ceess_probability, 6),
                "is_ceess_candidate": "yes" if is_candidate else "no",
                **probability_columns,
                **rawscore_columns,
            }
        )

    prediction_rows_sorted = sorted(
        prediction_rows,
        key=lambda row: (
            0 if row["is_ceess_candidate"] == "yes" else 1,
            -float(row["esm_ceess_probability"]),
            str(row["sequence_id"]),
        )
    )

    metrics_path = write_tsv(metrics_rows, output_root / "ceess_model_metrics.tsv")
    confusion_path = write_tsv(confusion_rows, output_root / "ceess_model_confusion_matrix.tsv")
    group_confusion_path = write_tsv(
        [
            {
                "true_ceess_group": label,
                "pred_CeeSs": int(row[0]),
                "pred_non_CeeSs": int(row[1]),
            }
            for label, row in zip(["CeeSs", "non-CeeSs"], group_confusion)
        ],
        output_root / "ceess_group_confusion_matrix.tsv",
    )
    projection_path = write_tsv(projection_rows, output_root / "ceess_projection.tsv")
    predictions_path = write_tsv(prediction_rows_sorted, output_root / "ceess_predictions.tsv")
    candidates_path = write_tsv(
        [row for row in prediction_rows_sorted if row["is_ceess_candidate"] == "yes"],
        output_root / "ceess_candidates.tsv",
    )
    candidate_fasta_path = write_fasta(ceess_candidate_records, output_root / "ceess_candidates.fasta")
    subtype_tsv_dir = ensure_directory(output_root / "type_score_hits")
    subtype_fasta_dir = ensure_directory(output_root / "type_score_fastas")
    candidate_record_by_id = {record.id: record for record in candidate_records}
    for label in classes:
        probability_column = _type_probability_column(label)
        output_stem = _type_output_stem(label)
        subtype_rows = []
        subtype_fasta_records: list[FastaRecord] = []
        for row in prediction_rows_sorted:
            try:
                probability_value = float(row.get(probability_column, "") or 0.0)
            except (TypeError, ValueError):
                probability_value = 0.0
            if probability_value <= 0.95:
                continue
            subtype_rows.append(dict(row))
            record = candidate_record_by_id.get(str(row["sequence_id"]))
            if record is not None:
                subtype_fasta_records.append(record.clone())
        write_tsv(subtype_rows, subtype_tsv_dir / f"{output_stem}.tsv")
        write_fasta(subtype_fasta_records, subtype_fasta_dir / f"{output_stem}.fasta")
    embedding_path = _render_ceess_candidate_svg(
        train_records,
        train_labels,
        train_coords,
        prediction_rows,
        query_coords,
        output_root / "ceess_embedding.svg",
        projection_method=projection_method,
    )

    return CeessPredictionResult(
        output_paths={
            "ceess_embedding_cache": output_root / "ceess_embedding_cache.npz",
            "ceess_barlow_representations": output_root / "ceess_barlow_representations.npz",
            "ceess_barlow_training": output_root / "ceess_barlow_training.tsv",
            "ceess_barlow_checkpoint": output_root / "ceess_barlow_checkpoint.pt",
            "ceess_classifier_checkpoint": output_root / "ceess_classifier_checkpoint.pt",
            "ceess_model_metrics": metrics_path,
            "ceess_model_confusion_matrix": confusion_path,
            "ceess_group_confusion_matrix": group_confusion_path,
            "ceess_projection": projection_path,
            "ceess_predictions": predictions_path,
            "ceess_candidates": candidates_path,
            "ceess_candidates_fasta": candidate_fasta_path,
            "type_score_hits": subtype_tsv_dir,
            "type_score_fastas": subtype_fasta_dir,
            "ceess_embedding_svg": embedding_path,
        },
        prediction_rows=prediction_rows,
        candidate_ids=[row["sequence_id"] for row in prediction_rows_sorted if row["is_ceess_candidate"] == "yes"],
    )
