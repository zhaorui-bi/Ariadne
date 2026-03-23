"""ESM embedding workflow for coral TPS type separation from spreadsheet data.

This module is intentionally kept separate from the main four-stage Ariadne
pipeline because it targets a supervised analysis setting: a curated Excel file
containing known coral TPS proteins and their type labels. The typical use case
is to embed these proteins with ESM2, then evaluate how well the resulting
representation separates product types such as CeeSs-related classes and other
TPS groups.
"""

from __future__ import annotations

import logging
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Optional, Union

import numpy as np
from sklearn.base import clone
from sklearn.decomposition import PCA
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, classification_report, confusion_matrix, f1_score
from sklearn.model_selection import StratifiedKFold, cross_val_predict
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler

from ariadne.fasta_utils import FastaRecord, clean_sequence, ensure_directory, write_fasta, write_tsv

PathLike = Union[str, Path]

logger = logging.getLogger(__name__)

DEFAULT_ESM_MODEL_NAME = "facebook/esm2_t33_650M_UR50D"
ESM_MODEL_PRESETS = {
    "150M": "facebook/esm2_t30_150M_UR50D",
    "650M": "facebook/esm2_t33_650M_UR50D",
    # Keep a compatibility shorthand for project users who refer to the larger checkpoint as "504M".
    "504M": "facebook/esm2_t33_650M_UR50D",
    "3B": "facebook/esm2_t36_3B_UR50D",
    "15B": "facebook/esm2_t48_15B_UR50D",
}


def resolve_esm_model_name(model_name: str) -> str:
    """Resolve user-facing ESM presets to concrete Hugging Face model ids."""
    key = str(model_name).strip()
    preset = ESM_MODEL_PRESETS.get(key.upper())
    if preset is not None:
        if key.upper() == "504M":
            logger.info("Requested ESM preset '504M'; using supported checkpoint %s", preset)
        return preset
    return key


def esm_model_help_text() -> str:
    """Return concise CLI help text listing common supported ESM presets."""
    return (
        "ESM model preset or Hugging Face id. "
        "Common presets: 150M, 650M, 3B, 15B. "
        "Default: 650M. The shorthand 504M is accepted and mapped to the bundled large ESM2 checkpoint."
    )


@dataclass
class TPSTypeRecord:
    """One labeled protein record loaded from ``TPS.xlsx``."""

    name: str
    sequence: str
    label: str
    species: str


@dataclass
class CeessPredictionResult:
    """Outputs produced when scoring coral-like candidates for CeeSs."""

    output_paths: dict[str, Path]
    prediction_rows: list[dict[str, object]]
    candidate_ids: list[str]


def load_tps_xlsx(
    xlsx_path: PathLike,
    *,
    sheet_name: Optional[str] = None,
    name_column: str = "Name",
    sequence_column: str = "Protein",
    label_column: str = "Type",
    species_column: str = "Species",
) -> list[TPSTypeRecord]:
    """Load coral TPS records from an Excel workbook.

    The expected spreadsheet layout follows the current `TPS/TPS.xlsx`
    convention:
    - column 1: name
    - column 2: protein sequence
    - column 3: type label
    - column 4: species
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read TPS.xlsx files.") from exc

    workbook = load_workbook(Path(xlsx_path), read_only=True, data_only=True)
    target_sheet = sheet_name or workbook.sheetnames[0]
    worksheet = workbook[target_sheet]
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        raise ValueError(f"Worksheet '{target_sheet}' is empty in {xlsx_path}.")

    header_map = {str(value).strip(): index for index, value in enumerate(header) if value is not None}
    required = {
        "name": name_column,
        "sequence": sequence_column,
        "label": label_column,
    }
    for field, column_name in required.items():
        if column_name not in header_map:
            raise ValueError(
                f"Missing required column '{column_name}' in worksheet '{target_sheet}'. "
                f"Available columns: {list(header_map)}"
            )

    records: list[TPSTypeRecord] = []
    for row_index, row in enumerate(rows, start=2):
        name_value = row[header_map[name_column]]
        sequence_value = row[header_map[sequence_column]]
        label_value = row[header_map[label_column]]
        species_value = row[header_map[species_column]] if species_column in header_map else ""

        if not sequence_value or not label_value:
            continue
        sequence = clean_sequence(str(sequence_value), keep_gaps=False).replace("*", "")
        if not sequence:
            logger.warning("Skipped empty/invalid protein sequence at row %d in %s", row_index, xlsx_path)
            continue

        records.append(
            TPSTypeRecord(
                name=str(name_value).strip() if name_value else f"record_{row_index}",
                sequence=sequence,
                label=str(label_value).strip(),
                species=str(species_value).strip() if species_value else "",
            )
        )
    if not records:
        raise ValueError(f"No valid protein records were found in {xlsx_path}.")
    return records


def _load_esm_components(model_name: str, device: Optional[str] = None):
    """Load tokenizer, model, and runtime device lazily."""
    try:
        import torch
        from transformers import AutoTokenizer, EsmModel
    except ImportError as exc:
        raise RuntimeError(
            "ESM analysis requires optional dependencies. "
            "Install them with `pip install 'ariadne-tps[esm]'` or "
            "`pip install torch transformers`."
        ) from exc

    runtime_device = device or ("cuda" if torch.cuda.is_available() else "cpu")
    tokenizer = AutoTokenizer.from_pretrained(model_name)
    model = EsmModel.from_pretrained(model_name)
    model.eval()
    model.to(runtime_device)
    return torch, tokenizer, model, runtime_device


def compute_esm_embeddings(
    records: list[TPSTypeRecord],
    *,
    model_name: str = DEFAULT_ESM_MODEL_NAME,
    batch_size: int = 4,
    max_length: int = 2048,
    device: Optional[str] = None,
) -> np.ndarray:
    """Compute mean-pooled ESM2 embeddings for a set of protein sequences."""
    model_name = resolve_esm_model_name(model_name)
    torch, tokenizer, model, runtime_device = _load_esm_components(model_name, device=device)
    all_embeddings: list[np.ndarray] = []

    for start in range(0, len(records), batch_size):
        batch = records[start : start + batch_size]
        sequences = [record.sequence for record in batch]
        encoded = tokenizer(
            sequences,
            return_tensors="pt",
            padding=True,
            truncation=True,
            max_length=max_length,
        )
        encoded = {key: value.to(runtime_device) for key, value in encoded.items()}
        with torch.no_grad():
            outputs = model(**encoded)
        hidden = outputs.last_hidden_state
        attention_mask = encoded["attention_mask"]

        for row_index in range(hidden.shape[0]):
            token_count = int(attention_mask[row_index].sum().item())
            if token_count <= 2:
                pooled = hidden[row_index, :token_count].mean(dim=0)
            else:
                # Exclude BOS/EOS tokens when computing the sequence embedding.
                pooled = hidden[row_index, 1 : token_count - 1].mean(dim=0)
            all_embeddings.append(pooled.detach().cpu().numpy())

    return np.vstack(all_embeddings)


def _build_classifier(random_state: int = 0) -> Pipeline:
    """Construct the default small-sample type classifier."""
    return Pipeline(
        steps=[
            ("scaler", StandardScaler()),
            (
                "classifier",
                LogisticRegression(
                    max_iter=4000,
                    class_weight="balanced",
                    random_state=random_state,
                ),
            ),
        ]
    )


def _projection_coordinates(embeddings: np.ndarray, labels: list[str]) -> tuple[np.ndarray, str]:
    """Compute a 2D visualization projection, preferring supervised LDA."""
    unique_labels = sorted(set(labels))
    if len(unique_labels) >= 2:
        max_components = min(2, len(unique_labels) - 1)
        if max_components >= 1:
            try:
                lda = LinearDiscriminantAnalysis(n_components=max_components)
                coords = lda.fit_transform(embeddings, labels)
                if coords.shape[1] == 1:
                    coords = np.pad(coords, ((0, 0), (0, 1)))
                return coords[:, :2], "lda"
            except Exception as error:
                logger.warning("LDA projection failed; falling back to PCA: %s", error)
    coords = PCA(n_components=2, random_state=0).fit_transform(embeddings)
    return coords, "pca"


def _fit_projection_model(embeddings: np.ndarray, labels: list[str]) -> tuple[object, str]:
    """Fit a reusable 2D projection model for training and query embeddings."""
    unique_labels = sorted(set(labels))
    if len(unique_labels) >= 2:
        max_components = min(2, len(unique_labels) - 1)
        if max_components >= 1:
            try:
                lda = LinearDiscriminantAnalysis(n_components=max_components)
                lda.fit(embeddings, labels)
                return lda, "lda"
            except Exception as error:
                logger.warning("LDA projection fit failed; falling back to PCA: %s", error)
    pca = PCA(n_components=2, random_state=0)
    pca.fit(embeddings)
    return pca, "pca"


def _transform_projection(model: object, embeddings: np.ndarray) -> np.ndarray:
    """Project embeddings with a fitted LDA/PCA model."""
    coords = np.asarray(model.transform(embeddings))
    if coords.ndim == 1:
        coords = coords[:, None]
    if coords.shape[1] == 1:
        coords = np.pad(coords, ((0, 0), (0, 1)))
    return coords[:, :2]


def _cross_validated_metrics(
    embeddings: np.ndarray,
    labels: list[str],
    *,
    cv_folds: int,
    random_state: int,
) -> tuple[np.ndarray, np.ndarray, list[dict[str, object]], list[dict[str, object]], int]:
    """Compute cross-validated predictions plus metric and confusion tables."""
    class_counts = Counter(labels)
    min_class_size = min(class_counts.values())
    n_splits = max(2, min(cv_folds, min_class_size))
    classifier = _build_classifier(random_state=random_state)
    splitter = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=random_state)
    cv_predictions = cross_val_predict(classifier, embeddings, labels, cv=splitter, method="predict")
    cv_probabilities = cross_val_predict(classifier, embeddings, labels, cv=splitter, method="predict_proba")

    label_order = sorted(class_counts)
    confusion = confusion_matrix(labels, cv_predictions, labels=label_order)
    confusion_rows = []
    for row_label, row_values in zip(label_order, confusion):
        row = {"true_type": row_label}
        for column_label, value in zip(label_order, row_values):
            row[column_label] = int(value)
        confusion_rows.append(row)

    report = classification_report(labels, cv_predictions, labels=label_order, output_dict=True, zero_division=0)
    metrics_rows = [
        {"metric": "num_sequences", "value": len(labels)},
        {"metric": "num_types", "value": len(class_counts)},
        {"metric": "cv_folds", "value": n_splits},
        {"metric": "cv_accuracy", "value": round(float(accuracy_score(labels, cv_predictions)), 6)},
        {"metric": "cv_macro_f1", "value": round(float(f1_score(labels, cv_predictions, average="macro")), 6)},
        {"metric": "cv_weighted_f1", "value": round(float(f1_score(labels, cv_predictions, average="weighted")), 6)},
    ]
    for label in label_order:
        label_report = report.get(label, {})
        metrics_rows.extend(
            [
                {"metric": f"{label}_precision", "value": round(float(label_report.get("precision", 0.0)), 6)},
                {"metric": f"{label}_recall", "value": round(float(label_report.get("recall", 0.0)), 6)},
                {"metric": f"{label}_f1", "value": round(float(label_report.get("f1-score", 0.0)), 6)},
                {"metric": f"{label}_support", "value": int(label_report.get("support", 0))},
            ]
        )
    return cv_predictions, cv_probabilities, metrics_rows, confusion_rows, n_splits


def _color_map(labels: list[str]) -> dict[str, str]:
    """Assign stable colors to type labels."""
    palette = [
        "#2563EB",
        "#0F766E",
        "#C2410C",
        "#7C3AED",
        "#DC2626",
        "#0891B2",
        "#65A30D",
        "#BE185D",
    ]
    return {label: palette[index % len(palette)] for index, label in enumerate(sorted(set(labels)))}


def _render_type_svg(
    records: list[TPSTypeRecord],
    labels: list[str],
    coords: np.ndarray,
    output_path: PathLike,
    *,
    projection_method: str,
) -> Path:
    """Render a clean 2D SVG showing ESM type separation."""
    target = Path(output_path)
    width = 1100
    height = 760
    left = 92
    top = 74
    right = 280
    bottom = 104
    plot_width = width - left - right
    plot_height = height - top - bottom
    plot_right = left + plot_width
    plot_bottom = top + plot_height

    xs = coords[:, 0]
    ys = coords[:, 1]
    min_x, max_x = float(xs.min()), float(xs.max())
    min_y, max_y = float(ys.min()), float(ys.max())
    pad_x = (max_x - min_x) * 0.08 or 0.5
    pad_y = (max_y - min_y) * 0.08 or 0.5
    min_x -= pad_x
    max_x += pad_x
    min_y -= pad_y
    max_y += pad_y

    def scale_x(value: float) -> float:
        return left + ((value - min_x) / max(max_x - min_x, 1e-9)) * plot_width

    def scale_y(value: float) -> float:
        return plot_bottom - ((value - min_y) / max(max_y - min_y, 1e-9)) * plot_height

    colors = _color_map(labels)
    svg: list[str] = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}" font-family="Helvetica, Arial, sans-serif">',
        f'<rect width="{width}" height="{height}" fill="white" />',
        f'<text x="{left}" y="38" font-size="28" font-weight="700" fill="#0B132B">Ariadne ESM Type Embedding</text>',
        f'<text x="{left}" y="60" font-size="12" fill="#64748B">Coral TPS proteins from TPS.xlsx projected with {projection_method.upper()} on top of mean-pooled ESM2 embeddings.</text>',
        f'<rect x="{left}" y="{top}" width="{plot_width}" height="{plot_height}" fill="#FFFFFF" stroke="#E2E8F0" stroke-width="1.2" />',
    ]

    for record, label, x_value, y_value in zip(records, labels, xs, ys):
        svg.append(
            f'<circle cx="{scale_x(float(x_value)):.2f}" cy="{scale_y(float(y_value)):.2f}" r="6.8" '
            f'fill="{colors[label]}" fill-opacity="0.9" stroke="white" stroke-width="1.2">'
            f'<title>{record.name} | {label}</title></circle>'
        )

    legend_x = plot_right + 34
    legend_y = top + 18
    svg.append(f'<text x="{legend_x}" y="{legend_y}" font-size="13" font-weight="700" fill="#0F172A">TPS types</text>')
    current_y = legend_y + 24
    for label in sorted(set(labels)):
        svg.append(f'<circle cx="{legend_x + 7}" cy="{current_y - 4}" r="6" fill="{colors[label]}" />')
        svg.append(f'<text x="{legend_x + 20}" y="{current_y}" font-size="12" fill="#334155">{label}</text>')
        current_y += 24

    svg.append(
        f'<text x="{left}" y="{height - 28}" font-size="10" fill="#94A3B8" font-style="italic">'
        f'Point labels are hidden to keep the figure readable. Inspect esm_projection.tsv and esm_predictions.tsv for per-record details.</text>'
    )
    svg.append("</svg>")
    target.write_text("".join(svg))
    return target


def _render_ceess_candidate_svg(
    train_records: list[TPSTypeRecord],
    train_labels: list[str],
    train_coords: np.ndarray,
    candidate_rows: list[dict[str, object]],
    candidate_coords: np.ndarray,
    output_path: PathLike,
    *,
    projection_method: str,
) -> Path:
    """Render a combined SVG with labeled TPS references and coral-like candidates."""
    target = Path(output_path)
    width = 1180
    height = 820
    left = 92
    top = 74
    right = 320
    bottom = 108
    plot_width = width - left - right
    plot_height = height - top - bottom
    plot_right = left + plot_width
    plot_bottom = top + plot_height

    all_coords = np.vstack([train_coords, candidate_coords]) if len(candidate_rows) else train_coords
    xs = all_coords[:, 0]
    ys = all_coords[:, 1]
    min_x, max_x = float(xs.min()), float(xs.max())
    min_y, max_y = float(ys.min()), float(ys.max())
    pad_x = (max_x - min_x) * 0.08 or 0.5
    pad_y = (max_y - min_y) * 0.08 or 0.5
    min_x -= pad_x
    max_x += pad_x
    min_y -= pad_y
    max_y += pad_y

    def scale_x(value: float) -> float:
        return left + ((value - min_x) / max(max_x - min_x, 1e-9)) * plot_width

    def scale_y(value: float) -> float:
        return plot_bottom - ((value - min_y) / max(max_y - min_y, 1e-9)) * plot_height

    colors = _color_map(train_labels)
    svg: list[str] = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}" font-family="Helvetica, Arial, sans-serif">',
        f'<rect width="{width}" height="{height}" fill="white" />',
        f'<text x="{left}" y="38" font-size="28" font-weight="700" fill="#0B132B">Ariadne CeeSs Candidate Embedding</text>',
        f'<text x="{left}" y="60" font-size="12" fill="#64748B">Coral TPS references from TPS.xlsx plus coral-like candidates projected with {projection_method.upper()} on mean-pooled ESM2 embeddings.</text>',
        f'<rect x="{left}" y="{top}" width="{plot_width}" height="{plot_height}" fill="#FFFFFF" stroke="#E2E8F0" stroke-width="1.2" />',
    ]

    for record, label, x_value, y_value in zip(train_records, train_labels, train_coords[:, 0], train_coords[:, 1]):
        svg.append(
            f'<circle cx="{scale_x(float(x_value)):.2f}" cy="{scale_y(float(y_value)):.2f}" r="5.8" '
            f'fill="{colors[label]}" fill-opacity="0.48" stroke="white" stroke-width="1.0">'
            f'<title>{record.name} | reference | {label}</title></circle>'
        )

    for row, x_value, y_value in zip(candidate_rows, candidate_coords[:, 0], candidate_coords[:, 1]):
        label = str(row.get("esm_ceess_label", row.get("esm_ceess_subtype", "")))
        fill = colors.get(label, "#111827")
        x = scale_x(float(x_value))
        y = scale_y(float(y_value))
        points = [
            f"{x:.2f},{y - 8.0:.2f}",
            f"{x + 8.0:.2f},{y:.2f}",
            f"{x:.2f},{y + 8.0:.2f}",
            f"{x - 8.0:.2f},{y:.2f}",
        ]
        stroke = "#111827" if row["is_ceess_candidate"] == "yes" else "#94A3B8"
        svg.append(
            f'<polygon points="{" ".join(points)}" fill="{fill}" fill-opacity="0.92" stroke="{stroke}" stroke-width="1.4">'
            f'<title>{row["sequence_id"]} | candidate | top={row["esm_type_prediction"]} | CeeSs={row["esm_ceess_probability"]}</title></polygon>'
        )

    legend_x = plot_right + 34
    legend_y = top + 18
    svg.append(f'<text x="{legend_x}" y="{legend_y}" font-size="13" font-weight="700" fill="#0F172A">Reference TPS types</text>')
    current_y = legend_y + 24
    for label in sorted(set(train_labels)):
        svg.append(f'<circle cx="{legend_x + 7}" cy="{current_y - 4}" r="6" fill="{colors[label]}" fill-opacity="0.68" />')
        svg.append(f'<text x="{legend_x + 20}" y="{current_y}" font-size="12" fill="#334155">{label}</text>')
        current_y += 24

    current_y += 14
    svg.append(f'<text x="{legend_x}" y="{current_y}" font-size="13" font-weight="700" fill="#0F172A">Coral-like candidates</text>')
    current_y += 24
    diamond = [
        f"{legend_x + 7:.2f},{current_y - 12:.2f}",
        f"{legend_x + 15:.2f},{current_y - 4:.2f}",
        f"{legend_x + 7:.2f},{current_y + 4:.2f}",
        f"{legend_x - 1:.2f},{current_y - 4:.2f}",
    ]
    svg.append(f'<polygon points="{" ".join(diamond)}" fill="#64748B" stroke="#111827" stroke-width="1.2" />')
    svg.append(f'<text x="{legend_x + 24}" y="{current_y}" font-size="12" fill="#334155">ESM-scored coral-like candidates</text>')
    current_y += 24
    svg.append(f'<text x="{legend_x}" y="{current_y}" font-size="11" fill="#64748B">Dark outline = final CeeSs candidate</text>')

    svg.append(
        f'<text x="{left}" y="{height - 28}" font-size="10" fill="#94A3B8" font-style="italic">'
        f'Candidate labels are hidden to keep the figure readable. Inspect ceess_predictions.tsv and ceess_candidates.tsv for per-record details.</text>'
    )
    svg.append("</svg>")
    target.write_text("".join(svg))
    return target


def analyze_tps_types_with_esm(
    xlsx_path: PathLike,
    output_dir: PathLike,
    *,
    sheet_name: Optional[str] = None,
    model_name: str = DEFAULT_ESM_MODEL_NAME,
    batch_size: int = 4,
    max_length: int = 2048,
    device: Optional[str] = None,
    cv_folds: int = 5,
    random_state: int = 0,
) -> dict[str, Path]:
    """Run a supervised ESM embedding workflow for coral TPS type separation."""
    model_name = resolve_esm_model_name(model_name)
    records = load_tps_xlsx(xlsx_path, sheet_name=sheet_name)
    labels = [record.label for record in records]
    class_counts = Counter(labels)
    logger.info("Loaded %d TPS proteins across %d types from %s", len(records), len(class_counts), xlsx_path)

    embeddings = compute_esm_embeddings(
        records,
        model_name=model_name,
        batch_size=batch_size,
        max_length=max_length,
        device=device,
    )

    cv_predictions, cv_probabilities, metrics_rows, confusion_rows, n_splits = _cross_validated_metrics(
        embeddings,
        labels,
        cv_folds=cv_folds,
        random_state=random_state,
    )

    classifier = _build_classifier(random_state=random_state)
    fitted = clone(classifier).fit(embeddings, labels)
    full_predictions = fitted.predict(embeddings)
    full_probabilities = fitted.predict_proba(embeddings)
    classes = list(fitted.named_steps["classifier"].classes_)

    coords, projection_method = _projection_coordinates(embeddings, labels)
    output_root = ensure_directory(output_dir)
    np.savez_compressed(
        output_root / "esm_embeddings.npz",
        embeddings=embeddings,
        labels=np.array(labels),
        names=np.array([record.name for record in records]),
    )

    projection_rows = []
    prediction_rows = []
    for index, record in enumerate(records):
        projection_rows.append(
            {
                "name": record.name,
                "type": record.label,
                "species": record.species,
                "length": len(record.sequence),
                "projection_method": projection_method,
                "axis1": round(float(coords[index, 0]), 6),
                "axis2": round(float(coords[index, 1]), 6),
            }
        )
        cv_pred = str(cv_predictions[index])
        full_pred = str(full_predictions[index])
        prediction_rows.append(
            {
                "name": record.name,
                "true_type": record.label,
                "cv_predicted_type": cv_pred,
                "cv_confidence": round(float(cv_probabilities[index].max()), 6),
                "full_predicted_type": full_pred,
                "full_confidence": round(float(full_probabilities[index].max()), 6),
                "species": record.species,
                "length": len(record.sequence),
            }
        )

    label_order = sorted(class_counts)
    metrics_rows.append({"metric": "esm_model", "value": model_name})

    centroid_rows = []
    for label in label_order:
        indices = [index for index, current in enumerate(labels) if current == label]
        centroid = embeddings[indices].mean(axis=0)
        centroid_rows.append(
            {
                "type": label,
                "count": len(indices),
                "axis1": round(float(coords[indices, 0].mean()), 6),
                "axis2": round(float(coords[indices, 1].mean()), 6),
                "embedding_norm": round(float(np.linalg.norm(centroid)), 6),
            }
        )

    projection_path = write_tsv(projection_rows, output_root / "esm_projection.tsv")
    predictions_path = write_tsv(prediction_rows, output_root / "esm_predictions.tsv")
    confusion_path = write_tsv(confusion_rows, output_root / "esm_confusion_matrix.tsv")
    metrics_path = write_tsv(metrics_rows, output_root / "esm_metrics.tsv")
    centroids_path = write_tsv(centroid_rows, output_root / "esm_type_centroids.tsv")
    svg_path = _render_type_svg(
        records,
        labels,
        coords,
        output_root / "esm_embedding.svg",
        projection_method=projection_method,
    )

    return {
        "embeddings": output_root / "esm_embeddings.npz",
        "projection": projection_path,
        "predictions": predictions_path,
        "confusion_matrix": confusion_path,
        "metrics": metrics_path,
        "type_centroids": centroids_path,
        "embedding_svg": svg_path,
    }


def classify_ceess_candidates_with_esm(
    candidate_records: list[FastaRecord],
    xlsx_path: PathLike,
    output_dir: PathLike,
    *,
    sheet_name: Optional[str] = None,
    model_name: str = DEFAULT_ESM_MODEL_NAME,
    batch_size: int = 4,
    max_length: int = 2048,
    device: Optional[str] = None,
    cv_folds: int = 5,
    random_state: int = 0,
    ceess_threshold: float = 0.5,
) -> CeessPredictionResult:
    """Score coral-like candidates with a binary CeeSs/non-CeeSs ESM classifier."""
    model_name = resolve_esm_model_name(model_name)
    if not candidate_records:
        raise ValueError("No coral-like candidate records were provided for CeeSs scoring.")

    train_records = load_tps_xlsx(xlsx_path, sheet_name=sheet_name)
    train_labels = [record.label for record in train_records]
    train_binary_labels = [
        "CeeSs" if label in {"cembrene A", "cembrene B"} else "non-CeeSs"
        for label in train_labels
    ]
    query_type_records = [
        TPSTypeRecord(
            name=record.id,
            sequence=record.sequence,
            label="candidate",
            species=record.metadata.get("source", "candidate"),
        )
        for record in candidate_records
    ]
    combined_records = train_records + query_type_records
    combined_embeddings = compute_esm_embeddings(
        combined_records,
        model_name=model_name,
        batch_size=batch_size,
        max_length=max_length,
        device=device,
    )
    train_embeddings = combined_embeddings[: len(train_records)]
    query_embeddings = combined_embeddings[len(train_records) :]

    cv_predictions, _, metrics_rows, confusion_rows, _ = _cross_validated_metrics(
        train_embeddings,
        train_binary_labels,
        cv_folds=cv_folds,
        random_state=random_state,
    )
    metrics_rows.extend(
        [
            {"metric": "esm_model", "value": model_name},
            {"metric": "ceess_threshold", "value": round(float(ceess_threshold), 6)},
            {"metric": "num_ceess_references", "value": sum(label == "CeeSs" for label in train_binary_labels)},
            {"metric": "num_non_ceess_references", "value": sum(label == "non-CeeSs" for label in train_binary_labels)},
            {"metric": "num_coral_like_candidates", "value": len(candidate_records)},
        ]
    )

    classifier = _build_classifier(random_state=random_state)
    fitted = clone(classifier).fit(train_embeddings, train_binary_labels)
    query_probabilities = fitted.predict_proba(query_embeddings)
    query_predictions = fitted.predict(query_embeddings)
    classes = list(fitted.named_steps["classifier"].classes_)
    class_to_index = {label: index for index, label in enumerate(classes)}

    projection_model, projection_method = _fit_projection_model(train_embeddings, train_binary_labels)
    train_coords = _transform_projection(projection_model, train_embeddings)
    query_coords = _transform_projection(projection_model, query_embeddings)

    output_root = ensure_directory(output_dir)
    projection_rows = []
    prediction_rows = []
    ceess_candidate_records: list[FastaRecord] = []
    for train_record, coords in zip(train_records, train_coords):
        binary_label = "CeeSs" if train_record.label in {"cembrene A", "cembrene B"} else "non-CeeSs"
        projection_rows.append(
            {
                "sequence_id": train_record.name,
                "role": "reference",
                "type": binary_label,
                "species": train_record.species,
                "projection_method": projection_method,
                "axis1": round(float(coords[0]), 6),
                "axis2": round(float(coords[1]), 6),
            }
        )

    for record, coords, probabilities, top_prediction in zip(candidate_records, query_coords, query_probabilities, query_predictions):
        ceess_probability = float(probabilities[class_to_index["CeeSs"]]) if "CeeSs" in class_to_index else 0.0
        non_ceess_probability = float(probabilities[class_to_index["non-CeeSs"]]) if "non-CeeSs" in class_to_index else (1.0 - ceess_probability)
        is_candidate = ceess_probability >= ceess_threshold
        if is_candidate:
            ceess_candidate_records.append(record.clone())
        projection_rows.append(
            {
                "sequence_id": record.id,
                "role": "candidate",
                "type": str(top_prediction),
                "species": record.metadata.get("source", "candidate"),
                "projection_method": projection_method,
                "axis1": round(float(coords[0]), 6),
                "axis2": round(float(coords[1]), 6),
            }
        )
        prediction_rows.append(
            {
                "sequence_id": record.id,
                "esm_type_prediction": str(top_prediction),
                "esm_ceess_label": str(top_prediction),
                "esm_ceess_subtype": "",
                "esm_ceess_probability": round(ceess_probability, 6),
                "esm_non_ceess_probability": round(non_ceess_probability, 6),
                "esm_cembrene_a_probability": "",
                "esm_cembrene_b_probability": "",
                "is_ceess_candidate": "yes" if is_candidate else "no",
            }
        )

    prediction_rows_sorted = sorted(
        prediction_rows,
        key=lambda row: (
            0 if row["is_ceess_candidate"] == "yes" else 1,
            -float(row["esm_ceess_probability"]),
            str(row["sequence_id"]),
        )
    )

    metrics_path = write_tsv(metrics_rows, output_root / "ceess_model_metrics.tsv")
    confusion_path = write_tsv(confusion_rows, output_root / "ceess_model_confusion_matrix.tsv")
    projection_path = write_tsv(projection_rows, output_root / "ceess_projection.tsv")
    predictions_path = write_tsv(prediction_rows_sorted, output_root / "ceess_predictions.tsv")
    candidates_path = write_tsv(
        [row for row in prediction_rows_sorted if row["is_ceess_candidate"] == "yes"],
        output_root / "ceess_candidates.tsv",
    )
    candidate_fasta_path = write_fasta(ceess_candidate_records, output_root / "ceess_candidates.fasta")
    embedding_path = _render_ceess_candidate_svg(
        train_records,
        train_binary_labels,
        train_coords,
        prediction_rows,
        query_coords,
        output_root / "ceess_embedding.svg",
        projection_method=projection_method,
    )

    return CeessPredictionResult(
        output_paths={
            "ceess_model_metrics": metrics_path,
            "ceess_model_confusion_matrix": confusion_path,
            "ceess_projection": projection_path,
            "ceess_predictions": predictions_path,
            "ceess_candidates": candidates_path,
            "ceess_candidates_fasta": candidate_fasta_path,
            "ceess_embedding_svg": embedding_path,
        },
        prediction_rows=prediction_rows,
        candidate_ids=[row["sequence_id"] for row in prediction_rows_sorted if row["is_ceess_candidate"] == "yes"],
    )
