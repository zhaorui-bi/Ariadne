from __future__ import annotations

from pathlib import Path
from typing import Optional, Union

from ariadne.fasta_utils import FastaRecord, ensure_directory, read_fasta, slugify, ungap, write_fasta, write_tsv

PathLike = Union[str, Path]


def _prepare_record(record: FastaRecord, *, source: str, extra: Optional[dict[str, str]] = None) -> FastaRecord:
    prepared = record.clone(sequence=ungap(record.sequence).replace("*", ""))
    prepared.metadata["source"] = source
    prepared.metadata["header"] = prepared.header
    prepared.metadata["label"] = "cembrene" if "cembrene" in prepared.header.lower() else "other"
    prepared.metadata["is_cembrene"] = "yes" if "cembrene" in prepared.header.lower() else "no"
    if extra:
        prepared.metadata.update(extra)
    return prepared


def prepare_coral_reference(
    input_fasta: PathLike,
    output_dir: PathLike,
    *,
    filename: str = "coral.fasta",
    limit: Optional[int] = None,
) -> tuple[Path, list[FastaRecord]]:
    destination = ensure_directory(output_dir)
    records = [_prepare_record(record, source="coral") for record in read_fasta(input_fasta, keep_gaps=True)]
    if limit is not None:
        records = records[:limit]
    output_path = destination / filename
    write_fasta(records, output_path)
    return output_path, records


def prepare_insect_reference(
    input_xlsx: PathLike,
    output_dir: PathLike,
    *,
    filename: str = "insect.fasta",
    sheet_name: str = "Protein Science",
    limit: Optional[int] = None,
) -> tuple[Path, list[FastaRecord]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to parse the insect reference workbook.") from exc

    workbook = openpyxl.load_workbook(Path(input_xlsx), read_only=True, data_only=True)
    sheet = workbook[sheet_name]

    header_map: dict[str, int] = {}
    header_row_index = None
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [value.strip() if isinstance(value, str) else value for value in row]
        if "Sequence ID" in values and "Sequence" in values:
            header_map = {str(value): index for index, value in enumerate(values) if value}
            header_row_index = row_index
            break
    if header_row_index is None:
        raise ValueError(f"Could not find the sequence table header in sheet '{sheet_name}'.")

    records: list[FastaRecord] = []
    for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        sequence = row[header_map["Sequence"]]
        if not sequence:
            continue
        sequence_id = row[header_map.get("Accession", -1)] or row[header_map["Sequence ID"]]
        species = row[header_map.get("Species", -1)] or "unknown_species"
        clade = row[header_map.get("Clade", -1)] or "unknown_clade"
        original_id = row[header_map["Sequence ID"]]
        accession = str(sequence_id).strip()
        header = f"{accession} {species} [{clade}]"
        record = FastaRecord(header=header, sequence=str(sequence).strip())
        record.metadata["source"] = "insect"
        record.metadata["header"] = header
        record.metadata["original_id"] = str(original_id).strip()
        record.metadata["species"] = str(species).strip()
        record.metadata["clade"] = str(clade).strip()
        record.metadata["label"] = str(clade).strip()
        record.metadata["is_cembrene"] = "no"
        records.append(record)
        if limit is not None and len(records) >= limit:
            break

    destination = ensure_directory(output_dir)
    output_path = destination / filename
    write_fasta(records, output_path)
    return output_path, records


def prepare_extra_reference(input_fasta: PathLike, output_dir: PathLike, *, source: str) -> tuple[Path, list[FastaRecord]]:
    destination = ensure_directory(output_dir)
    filename = f"{slugify(source)}.fasta"
    records = [_prepare_record(record, source=source) for record in read_fasta(input_fasta, keep_gaps=True)]
    output_path = destination / filename
    write_fasta(records, output_path)
    return output_path, records


def write_reference_metadata(records: list[FastaRecord], output_dir: PathLike, *, filename: str = "metadata.tsv") -> Path:
    rows: list[dict[str, object]] = []
    for record in records:
        row = {"sequence_id": record.id, "header": record.header}
        row.update(record.metadata)
        rows.append(row)
    return write_tsv(rows, ensure_directory(output_dir) / filename)


def load_reference_records(reference_dir: PathLike) -> list[FastaRecord]:
    directory = Path(reference_dir)
    metadata_map: dict[str, dict[str, str]] = {}
    metadata_path = directory / "metadata.tsv"
    if metadata_path.exists():
        import csv

        with metadata_path.open() as handle:
            reader = csv.DictReader(handle, delimiter="\t")
            for row in reader:
                metadata_map[row["sequence_id"]] = {key: value for key, value in row.items() if value}

    records: list[FastaRecord] = []
    for fasta_path in sorted(directory.glob("*.fa*")):
        source_name = fasta_path.stem.split(".", 1)[0]
        for record in read_fasta(fasta_path):
            record.metadata.setdefault("source", source_name)
            record.metadata.setdefault("header", record.header)
            if record.id in metadata_map:
                record.metadata.update(metadata_map[record.id])
            records.append(record)
    return records
